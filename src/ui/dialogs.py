"""对话框组件"""
import os
import re
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QComboBox, QLineEdit, QTextEdit, QSpinBox, QDateEdit,
    QDialogButtonBox, QMessageBox, QAbstractItemView, QHeaderView,
    QCheckBox,
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor

from src.models.database import (
    get_connection, ship_quote, add_quote, update_quote, add_payment,
    get_payments, get_operation_logs, get_batches, get_batch_remaining,
    get_customer_statement, get_all_customers, search_customers,
    add_customer, get_all_suppliers, get_supplier_payable,
    get_quote_by_id, get_all_products, search_products,
    add_supplier,
)
from src.utils.shipment_flow import parse_sn_input, validate_sn_list, check_sn_duplicates


from src.ui.utils import _validate_date


def _parse_tax_rate(combo):
    """解析税率下拉框的值，支持预设选项和手动输入"""
    data = combo.currentData()
    if data is not None:
        return data if data > 0 else None
    text = combo.currentText().strip().replace("%", "").strip()
    if not text:
        return None
    try:
        val = float(text)
        return val / 100.0 if val > 1 else val
    except ValueError:
        return None


# ============================================================
# 出库对话框
# ============================================================
class ShipmentDialog(QDialog):
    def __init__(self, parent=None, quote=None, batches=None):
        super().__init__(parent)
        self.quote = quote
        self.batches = batches or []
        self.setWindowTitle("确认出库")
        self.setMinimumWidth(450)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        info_text = f"机型: {quote.get('series','')} {quote.get('cpu','')}\n客户: {quote.get('customer_name','')}\n数量: {quote.get('quote_quantity',1)} 台"
        self.info_label = QLabel(info_text)
        self.info_label.setObjectName("dialogInfoLabel")
        layout.addRow(self.info_label)

        self.batch_combo = QComboBox()
        for b in self.batches:
            self.batch_combo.addItem(
                f"批次#{b['id']} | 购入¥{b['purchase_price']:.0f} | 剩余{b['remaining']}台 | {b.get('date','')}",
                b["id"]
            )
        self.batch_combo.currentIndexChanged.connect(self._update_remaining)
        layout.addRow("扣减批次:", self.batch_combo)

        self.remaining_label = QLabel()
        self._update_remaining()
        layout.addRow("批次剩余:", self.remaining_label)

        self.sn_edit = QTextEdit()
        self.sn_edit.setPlaceholderText("条码枪扫码输入，每扫一条自动换行。也可手动输入，支持逗号/空格分隔。")
        self.sn_edit.setMaximumHeight(100)
        layout.addRow("出库SN:", self.sn_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("出库备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

        # 条码枪支持：拦截回车键，不让它提交对话框
        from PyQt6.QtCore import QEvent
        self.sn_edit.installEventFilter(self)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent, Qt
        if obj == self.sn_edit and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self.sn_edit.insertPlainText("\n")
                return True
        return super().eventFilter(obj, event)

    def _update_remaining(self):
        idx = self.batch_combo.currentIndex()
        if 0 <= idx < len(self.batches):
            remaining = self.batches[idx].get("remaining", 0)
            self.remaining_label.setText(f"{remaining} 台")
            color = "#D32F2F" if remaining < (self.quote.get("quote_quantity", 1) or 1) else "#388E3C"
            self.remaining_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def _validate_and_accept(self):
        quantity = self.quote.get("quote_quantity", 1) or 1
        idx = self.batch_combo.currentIndex()
        if idx < 0 or idx >= len(self.batches):
            QMessageBox.warning(self, "提示", "请选择批次")
            return
        remaining = self.batches[idx].get("remaining", 0)
        if remaining < quantity:
            QMessageBox.warning(self, "提示", f"批次剩余不足！需要 {quantity} 台，剩余 {remaining} 台")
            return
        raw_sn = self.sn_edit.toPlainText().strip()
        if raw_sn:
            sn_list = parse_sn_input(raw_sn)
            sn_pattern = re.compile(r"^[A-Za-z0-9\-]{4,20}$")
            for sn in sn_list:
                if not sn_pattern.match(sn):
                    QMessageBox.warning(self, "SN格式错误",
                        f"序列号格式不正确: {sn}\n"
                        "SN应为4-20位字母/数字/连字符，不含特殊符号")
                    return
        self.accept()

    def get_data(self):
        idx = self.batch_combo.currentIndex()
        raw_sn = self.sn_edit.toPlainText().strip()
        sn_list = parse_sn_input(raw_sn)
        return {
            "batch_id": self.batch_combo.currentData() if idx >= 0 else None,
            "sn_list": ",".join(sn_list),
            "sn_count": len(sn_list),
            "remark": self.remark_edit.text().strip(),
        }


# ============================================================
# 收款/付款对话框
# ============================================================
class PaymentDialog(QDialog):
    def __init__(self, parent=None, title="收款", pay_type="receivable", quote=None, preview_pending=None):
        super().__init__(parent)
        self.pay_type = pay_type
        self.quote = quote
        self.preview_pending = preview_pending
        self.setWindowTitle(title)
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        if quote:
            total_amount = (quote.get("quote_price", 0) or 0) * (quote.get("quote_quantity", 1) or 1)
            received = quote.get("received_amount", 0) or 0
            remaining = total_amount - received
            info = f"客户: {quote.get('customer_name','')} | 总金额: ¥{total_amount:.0f} | 已收: ¥{received:.0f} | 待收: ¥{remaining:.0f}"
            info_label = QLabel(info)
            info_label.setObjectName("dialogInfoLabel")
            layout.addRow(info_label)

        if preview_pending:
            info_label = QLabel(f"当前待收金额: ¥{preview_pending:.0f}")
            info_label.setObjectName("dangerSummaryLabel")
            layout.addRow(info_label)

        self.amount_spin = QSpinBox()
        self.amount_spin.setRange(1, 999999)
        self.amount_spin.setPrefix("¥ ")
        self.amount_spin.setValue(0)
        layout.addRow("金额:", self.amount_spin)

        self.method_combo = QComboBox()
        self.method_combo.addItems(["微信", "支付宝", "转账", "现金"])
        layout.addRow("方式:", self.method_combo)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        layout.addRow("日期:", self.date_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _validate_and_accept(self):
        if self.amount_spin.value() <= 0:
            QMessageBox.warning(self, "提示", "金额必须大于0")
            return
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        valid, msg = _validate_date(date_str)
        if not valid:
            QMessageBox.warning(self, "日期错误", msg)
            return
        self.accept()

    def get_data(self):
        return {
            "amount": self.amount_spin.value(),
            "method": self.method_combo.currentText(),
            "pay_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
        }


# ============================================================
# 收付款记录编辑对话框
# ============================================================
class PaymentEditDialog(QDialog):
    def __init__(self, parent=None, payment=None):
        super().__init__(parent)
        self.payment = payment
        self.setWindowTitle("编辑收付款记录")
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        # 显示原记录信息
        type_text = "收款" if payment.get("type") == "receivable" else "付款"
        obj_name = payment.get("customer_name", "") or payment.get("supplier_name", "")
        info = f"类型: {type_text} | 关联对象: {obj_name}"
        info_label = QLabel(info)
        info_label.setObjectName("dialogInfoLabel")
        layout.addRow(info_label)

        self.amount_spin = QSpinBox()
        self.amount_spin.setRange(1, 999999)
        self.amount_spin.setPrefix("¥ ")
        self.amount_spin.setValue(int(payment.get("amount", 0)))
        layout.addRow("金额:", self.amount_spin)

        self.method_combo = QComboBox()
        self.method_combo.addItems(["微信", "支付宝", "转账", "现金"])
        # 设置当前方式
        current_method = payment.get("method", "")
        idx = self.method_combo.findText(current_method)
        if idx >= 0:
            self.method_combo.setCurrentIndex(idx)
        layout.addRow("方式:", self.method_combo)

        self.date_edit = QDateEdit()
        if payment.get("pay_date"):
            qdate = QDate.fromString(payment.get("pay_date"), "yyyy-MM-dd")
            if qdate.isValid():
                self.date_edit.setDate(qdate)
        else:
            self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        layout.addRow("日期:", self.date_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setText(payment.get("remark", "") or "")
        self.remark_edit.setPlaceholderText("备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "amount": self.amount_spin.value(),
            "method": self.method_combo.currentText(),
            "pay_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
        }


# ============================================================
# 客户对账单对话框
# ============================================================
class StatementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("客户对账单")
        self.setMinimumSize(800, 550)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("客户:"))

        self.customer_combo = QComboBox()
        self.customer_combo.setEditable(True)
        self.customer_combo.setPlaceholderText("选择或输入客户...")
        customers = get_all_customers()
        for c in customers:
            self.customer_combo.addItem(c["name"], c["id"])
        top_row.addWidget(self.customer_combo)

        top_row.addWidget(QLabel("开始:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        top_row.addWidget(self.date_from)

        top_row.addWidget(QLabel("结束:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        top_row.addWidget(self.date_to)

        self.query_btn = QPushButton("查询")
        self.query_btn.setObjectName("ghostBtn")
        self.query_btn.clicked.connect(self._query)
        top_row.addWidget(self.query_btn)

        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setObjectName("successBtn")
        self.export_btn.clicked.connect(self._export_excel)
        top_row.addWidget(self.export_btn)

        layout.addLayout(top_row)

        self.summary_label = QLabel()
        self.summary_label.setObjectName("summaryLabel")
        layout.addWidget(self.summary_label)

        self.statement_table = QTableWidget()
        self.statement_table.setAlternatingRowColors(True)
        self.statement_table.setColumnCount(11)
        self.statement_table.setHorizontalHeaderLabels(
            ["日期", "机型", "CPU", "数量", "购入价", "报价", "毛利", "状态", "已收款", "待收款", "备注"]
        )
        self.statement_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.statement_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.statement_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.statement_table)

    def _query(self):
        customer_name = self.customer_combo.currentText().strip()
        customer_id = self.customer_combo.currentData()
        if not customer_id and customer_name:
            matched = search_customers(customer_name)
            if matched:
                customer_id = matched[0]["id"]
            else:
                QMessageBox.warning(self, "提示", "未找到该客户")
                return
        if not customer_id:
            QMessageBox.warning(self, "提示", "请选择客户")
            return

        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        self._records = get_customer_statement(customer_id, date_from, date_to)

        self.statement_table.setRowCount(len(self._records))
        total_cost = 0
        total_sale = 0
        total_received = 0
        for i, r in enumerate(self._records):
            purchase_price = r.get("purchase_price", 0) or 0
            quote_price = r.get("quote_price", 0) or 0
            quantity = r.get("quote_quantity", 1) or 1
            received = r.get("received_amount", 0) or 0
            total_amount = quote_price * quantity
            profit = (quote_price - purchase_price) * quantity
            pending = total_amount - received

            self.statement_table.setItem(i, 0, QTableWidgetItem(r.get("quote_date", "")))
            self.statement_table.setItem(i, 1, QTableWidgetItem(r.get("series", "")))
            self.statement_table.setItem(i, 2, QTableWidgetItem(r.get("cpu", "")))
            self.statement_table.setItem(i, 3, QTableWidgetItem(str(quantity)))
            self.statement_table.setItem(i, 4, QTableWidgetItem(f"¥{purchase_price:.0f}"))
            self.statement_table.setItem(i, 5, QTableWidgetItem(f"¥{quote_price:.0f}"))
            profit_item = QTableWidgetItem(f"¥{profit:.0f}" if profit >= 0 else f"-¥{-profit:.0f}")
            profit_item.setForeground(QColor("#388E3C") if profit >= 0 else QColor("#D32F2F"))
            self.statement_table.setItem(i, 6, profit_item)
            self.statement_table.setItem(i, 7, QTableWidgetItem(r.get("status", "")))
            self.statement_table.setItem(i, 8, QTableWidgetItem(f"¥{received:.0f}"))
            self.statement_table.setItem(i, 9, QTableWidgetItem(f"¥{pending:.0f}" if pending > 0 else "已结清"))
            merged_remark = " | ".join(filter(None, [r.get("batch_remark", "") or "", r.get("remark", "") or ""]))
            self.statement_table.setItem(i, 10, QTableWidgetItem(merged_remark))

            total_cost += purchase_price * quantity
            total_sale += total_amount
            total_received += received

        self.statement_table.resizeColumnsToContents()

        total_pending = total_sale - total_received
        self.summary_label.setText(
            f"客户: {customer_name} | 总购入: ¥{total_cost:.0f} | 总金额: ¥{total_sale:.0f} | "
            f"已收款: ¥{total_received:.0f} | 待收款: ¥{total_pending:.0f} | 总毛利: ¥{total_sale - total_cost:.0f}"
        )

    def _export_excel(self):
        if not hasattr(self, '_records') or not self._records:
            QMessageBox.warning(self, "提示", "请先查询数据")
            return
        customer_name = self.customer_combo.currentText().strip()
        output = export_statement_to_excel(self._records, customer_name)
        QMessageBox.information(self, "导出成功", f"对账单已导出:\n{output}")


def export_statement_to_excel(records, customer_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(desktop, f"对账单_{customer_name}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"

    headers = ["日期", "机型", "CPU", "数量", "购入价", "报价", "毛利", "状态", "已收款", "待收款", "备注"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_alignment = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    total_cost = 0
    total_sale = 0
    total_received = 0
    for row_idx, r in enumerate(records, 2):
        purchase_price = r.get("purchase_price", 0) or 0
        quote_price = r.get("quote_price", 0) or 0
        quantity = r.get("quote_quantity", 1) or 1
        received = r.get("received_amount", 0) or 0
        total_amount = quote_price * quantity
        profit = (quote_price - purchase_price) * quantity
        pending = total_amount - received
        merged_remark = " | ".join(filter(None, [r.get("batch_remark", "") or "", r.get("remark", "") or ""]))

        row_data = [
            r.get("quote_date", ""), r.get("series", ""), r.get("cpu", ""),
            quantity, purchase_price, quote_price, profit, r.get("status", ""),
            received, pending if pending > 0 else "已结清", merged_remark,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        total_cost += purchase_price * quantity
        total_sale += total_amount
        total_received += received

    summary_row = len(records) + 2
    summary_data = [
        "合计", "", "", "",
        round(total_cost, 2), round(total_sale, 2),
        round(total_sale - total_cost, 2), "",
        round(total_received, 2), round(total_sale - total_received, 2), "",
    ]
    for col_idx, val in enumerate(summary_data, 1):
        cell = ws.cell(row=summary_row, column=col_idx, value=val)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border

    col_widths = [12, 16, 14, 8, 10, 10, 10, 10, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# 编辑机型对话框
# ============================================================
class ProductEditDialog(QDialog):
    def __init__(self, parent=None, product=None):
        super().__init__(parent)
        self.product = product  # dict or None
        self.setWindowTitle("编辑机型" if product else "新增机型")
        self.setMinimumWidth(450)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.series_edit = QLineEdit()
        self.cpu_edit = QLineEdit()
        self.ram_edit = QLineEdit()
        self.storage_edit = QLineEdit()
        self.gpu_edit = QLineEdit()
        self.screen_edit = QLineEdit()
        self.note_edit = QLineEdit()

        if product:
            self.series_edit.setText(product.get("series", ""))
            self.cpu_edit.setText(product.get("cpu", ""))
            self.ram_edit.setText(product.get("ram", ""))
            self.storage_edit.setText(product.get("storage", ""))
            self.gpu_edit.setText(product.get("gpu", ""))
            self.screen_edit.setText(product.get("screen", ""))
            self.note_edit.setText(product.get("note", ""))

        layout.addRow("系列:", self.series_edit)
        layout.addRow("CPU:", self.cpu_edit)
        layout.addRow("内存:", self.ram_edit)
        layout.addRow("硬盘:", self.storage_edit)
        layout.addRow("显卡:", self.gpu_edit)
        layout.addRow("屏幕:", self.screen_edit)
        layout.addRow("备注:", self.note_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _validate_and_accept(self):
        if not self.series_edit.text().strip():
            QMessageBox.warning(self, "提示", "系列不能为空")
            return
        self.accept()

    def get_data(self):
        return {
            "series": self.series_edit.text().strip(),
            "cpu": self.cpu_edit.text().strip(),
            "ram": self.ram_edit.text().strip(),
            "storage": self.storage_edit.text().strip(),
            "gpu": self.gpu_edit.text().strip(),
            "screen": self.screen_edit.text().strip(),
            "note": self.note_edit.text().strip(),
        }


# ============================================================
# 新增批次库存对话框
# ============================================================
class BatchDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新增库存批次")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.price_spin = QSpinBox()
        self.price_spin.setRange(0, 999999)
        self.price_spin.setPrefix("¥ ")
        self.price_spin.setValue(0)

        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 9999)
        self.quantity_spin.setValue(1)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setEditable(True)
        self.supplier_combo.setPlaceholderText("输入上游名称或选择...")
        self._load_suppliers()

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("如: 含税/未税")

        self.sn_edit = QLineEdit()
        self.sn_edit.setPlaceholderText("序列号，多台用逗号分隔（选填）")

        layout.addRow("购入价:", self.price_spin)
        layout.addRow("数量:", self.quantity_spin)
        layout.addRow("入库日期:", self.date_edit)
        layout.addRow("上游:", self.supplier_combo)
        layout.addRow("备注:", self.remark_edit)
        layout.addRow("序列号:", self.sn_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _load_suppliers(self):
        suppliers = get_all_suppliers()
        for s in suppliers:
            self.supplier_combo.addItem(s["name"], s["id"])

    def _validate_and_accept(self):
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        valid, msg = _validate_date(date_str)
        if not valid:
            QMessageBox.warning(self, "日期错误", msg)
            return
        self.accept()

    def get_data(self):
        supplier_name = self.supplier_combo.currentText().strip()
        supplier_id = self.supplier_combo.currentData()
        if not supplier_id and supplier_name:
            supplier_id = add_supplier(supplier_name)
        return {
            "price": self.price_spin.value(),
            "quantity": self.quantity_spin.value(),
            "date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
            "supplier_id": supplier_id,
            "sn_list": self.sn_edit.text().strip(),
        }


# ============================================================
# 客户编辑对话框
# ============================================================
class CustomerDialog(QDialog):
    def __init__(self, parent=None, customer=None):
        super().__init__(parent)
        self.customer = customer
        self.setWindowTitle("编辑客户" if customer else "新增客户")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.name_edit = QLineEdit()
        self.wechat_edit = QLineEdit()
        self.qq_edit = QLineEdit()
        self.phone_edit = QLineEdit()
        self.note_edit = QLineEdit()

        self.tax_combo = QComboBox()
        self.tax_combo.setEditable(True)
        self.tax_combo.addItem("无", None)
        self.tax_combo.addItem("8%", 0.08)
        self.tax_combo.addItem("13%", 0.13)
        self.tax_combo.setCurrentIndex(0)

        if customer:
            self.name_edit.setText(customer.get("name", ""))
            self.wechat_edit.setText(customer.get("wechat", ""))
            self.qq_edit.setText(customer.get("qq", ""))
            self.phone_edit.setText(customer.get("phone", ""))
            self.note_edit.setText(customer.get("note", ""))
            tax_rate = customer.get("default_tax_rate")
            if tax_rate is not None:
                idx = self.tax_combo.findData(tax_rate)
                if idx >= 0:
                    self.tax_combo.setCurrentIndex(idx)
                else:
                    self.tax_combo.setEditText(f"{int(tax_rate * 100)}%")

        layout.addRow("名称:", self.name_edit)
        layout.addRow("微信:", self.wechat_edit)
        layout.addRow("QQ:", self.qq_edit)
        layout.addRow("电话:", self.phone_edit)
        layout.addRow("备注:", self.note_edit)

        tax_layout = QHBoxLayout()
        tax_layout.addWidget(QLabel("默认税率:"))
        tax_layout.addWidget(self.tax_combo)
        tax_layout.addStretch()
        layout.addRow(tax_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "name": self.name_edit.text().strip(),
            "wechat": self.wechat_edit.text().strip(),
            "qq": self.qq_edit.text().strip(),
            "phone": self.phone_edit.text().strip(),
            "note": self.note_edit.text().strip(),
            "default_tax_rate": _parse_tax_rate(self.tax_combo),
        }


# ============================================================
# 报价记录编辑对话框
# ============================================================
class QuoteEditDialog(QDialog):
    def __init__(self, parent=None, quote=None, batch_id=None, customer_id=None):
        super().__init__(parent)
        self.quote = quote
        self.setWindowTitle("编辑报价记录" if quote else "新增报价记录")
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.price_spin = QSpinBox()
        self.price_spin.setRange(0, 999999)
        self.price_spin.setPrefix("¥ ")
        self.price_spin.setValue(0)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.customer_combo = QComboBox()
        self.customer_combo.setEditable(True)
        self.customer_combo.setPlaceholderText("输入客户名称或选择...")
        self._load_customers()
        if customer_id:
            idx = self.customer_combo.findData(customer_id)
            if idx >= 0:
                self.customer_combo.setCurrentIndex(idx)

        self.remark_edit = QLineEdit()

        self.sn_edit = QLineEdit()
        self.sn_edit.setPlaceholderText("序列号，多台用逗号分隔（选填）")

        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 9999)
        self.quantity_spin.setValue(1)

        self.paid_combo = QComboBox()
        self.paid_combo.addItems(["否", "是"])

        # 价税覆盖层
        self.tax_combo = QComboBox()
        self.tax_combo.setEditable(True)
        self.tax_combo.addItem("无", None)
        self.tax_combo.addItem("8%", 0.08)
        self.tax_combo.addItem("13%", 0.13)
        self.tax_combo.setCurrentIndex(0)
        self.purchase_tax_check = QCheckBox("进价含税")
        self.quote_tax_check = QCheckBox("售价含税")

        if quote:
            self.price_spin.setValue(int(quote.get("quote_price", 0)))
            self.quantity_spin.setValue(quote.get("quote_quantity", 1))
            if quote.get("quote_date"):
                qdate = QDate.fromString(quote.get("quote_date"), "yyyy-MM-dd")
                if qdate.isValid():
                    self.date_edit.setDate(qdate)
            self.remark_edit.setText(quote.get("remark", ""))
            self.sn_edit.setText(quote.get("sn_list", "") or quote.get("batch_sn_list", ""))
            paid_value = quote.get("paid", "否")
            idx = self.paid_combo.findText(paid_value)
            if idx >= 0:
                self.paid_combo.setCurrentIndex(idx)
            tax_rate = quote.get("tax_rate")
            if tax_rate is not None:
                idx = self.tax_combo.findData(tax_rate)
                if idx >= 0:
                    self.tax_combo.setCurrentIndex(idx)
                else:
                    self.tax_combo.setEditText(f"{int(tax_rate * 100)}%")
            self.purchase_tax_check.setChecked(quote.get("purchase_tax_inclusive", 0) == 1)
            self.quote_tax_check.setChecked(quote.get("quote_tax_inclusive", 0) == 1)
        elif customer_id:
            # 新建报价时自动带入客户默认税率
            from src.models.database import get_customer_default_tax_rate
            default_tax = get_customer_default_tax_rate(customer_id)
            if default_tax is not None:
                idx = self.tax_combo.findData(default_tax)
                if idx >= 0:
                    self.tax_combo.setCurrentIndex(idx)
                else:
                    self.tax_combo.setEditText(f"{int(default_tax * 100)}%")

        layout.addRow("对外报价:", self.price_spin)
        layout.addRow("数量:", self.quantity_spin)
        layout.addRow("报价日期:", self.date_edit)
        layout.addRow("客户:", self.customer_combo)
        layout.addRow("备注:", self.remark_edit)
        layout.addRow("序列号:", self.sn_edit)
        layout.addRow("是否打款:", self.paid_combo)

        tax_layout = QHBoxLayout()
        tax_layout.addWidget(QLabel("税率:"))
        tax_layout.addWidget(self.tax_combo)
        tax_layout.addWidget(self.purchase_tax_check)
        tax_layout.addWidget(self.quote_tax_check)
        tax_layout.addStretch()
        layout.addRow(tax_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _load_customers(self):
        customers = get_all_customers()
        for c in customers:
            self.customer_combo.addItem(c["name"], c["id"])

    def _validate_and_accept(self):
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        valid, msg = _validate_date(date_str)
        if not valid:
            QMessageBox.warning(self, "日期错误", msg)
            return
        self.accept()

    def get_data(self):
        customer_name = self.customer_combo.currentText().strip()
        customer_id = self.customer_combo.currentData()
        if not customer_id and customer_name:
            customer_id = add_customer(customer_name)
        return {
            "quote_price": self.price_spin.value(),
            "quote_quantity": self.quantity_spin.value(),
            "quote_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "customer_id": customer_id,
            "remark": self.remark_edit.text().strip(),
            "sn_list": self.sn_edit.text().strip(),
            "paid": self.paid_combo.currentText(),
            "batch_id": self.quote.get("batch_id") if self.quote else None,
            "tax_rate": _parse_tax_rate(self.tax_combo),
            "purchase_tax_inclusive": 1 if self.purchase_tax_check.isChecked() else 0,
            "quote_tax_inclusive": 1 if self.quote_tax_check.isChecked() else 0,
        }


# ============================================================
# 操作日志对话框
# ============================================================
class OperationLogDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("最近操作记录")
        self.setMinimumSize(700, 400)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["时间", "操作", "对象", "描述"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSortingEnabled(True)

        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        refresh_btn = QPushButton("刷新")
        refresh_btn.setObjectName("ghostBtn")
        refresh_btn.clicked.connect(self.load_logs)
        btn_row.addWidget(refresh_btn)
        close_btn = QPushButton("关闭")
        close_btn.setObjectName("ghostBtn")
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self.load_logs()

    def load_logs(self):
        from src.models.database import get_operation_logs
        logs = get_operation_logs(limit=50)
        self.table.setRowCount(len(logs))
        for i, log in enumerate(logs):
            self.table.setItem(i, 0, QTableWidgetItem(log.get("created_at", "")))
            self.table.setItem(i, 1, QTableWidgetItem(log.get("operation", "")))
            self.table.setItem(i, 2, QTableWidgetItem(
                f"{log.get('table_name', '')}(ID:{log.get('record_id', '')})"
            ))
            self.table.setItem(i, 3, QTableWidgetItem(log.get("description", "")))
        self.table.resizeColumnsToContents()
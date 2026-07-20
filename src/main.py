"""
主窗口及各个 UI 面板
"""

import sys
import os
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QSplitter, QTabWidget, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QLabel, QComboBox, QTextEdit,
    QMessageBox, QFileDialog, QDialog, QFormLayout,
    QSpinBox, QDateEdit, QDialogButtonBox,
    QFrame, QHeaderView, QAbstractItemView, QCheckBox, QGroupBox,
    QGridLayout, QMenu, QMenuBar, QStatusBar,
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QAction, QClipboard, QColor

# 确保能找到 src 包
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from src.models.database import (
    init_db, add_product, update_product, delete_product,
    search_products, get_all_products,
    add_batch, get_batches, update_batch_remaining, delete_batch, get_total_remaining,
    add_customer, search_customers, get_all_customers,
    add_supplier, search_suppliers, get_all_suppliers,
    add_quote, update_quote, delete_quote, search_quotes, export_quotes, get_quote_by_id,
    update_quote_status, add_payment, _add_payment_raw, get_payments, get_customer_balance,
    get_supplier_payable, get_customer_statement, deduct_batch_remaining,
    delete_customer_cascade, delete_supplier_cascade,
    get_payment_by_id, get_all_payments_with_details, update_payment, delete_payment,
    get_connection, ship_quote, add_operation_log,
)
from src.utils.word_parser import parse_word_pricelist, preview_parse
from src.utils.image_gen import generate_quote_image, generate_single_quote_card, WATERMARK_TEXT
from src.utils.excel_export import export_quotes_to_excel
from src.utils.price_diff import save_snapshot, get_latest_snapshot, diff_snapshots, get_all_snapshots
from src.utils.follow_up import get_stale_quotes, format_reminder_text
from src.utils.monthly_report import get_monthly_report, format_report_text
from src.utils.shipment_flow import parse_sn_input, validate_sn, validate_sn_list, check_sn_duplicates, generate_shipment_receipt
from src.utils.quote_assist import get_quote_history, suggest_price, get_customer_price_history
from src.utils.remote_diagnose import search_diagnose, get_diagnose_tree, get_all_diagnose_keys, generate_diagnose_report


# ============================================================
# 出库对话框
# ============================================================
class ShipmentDialog(QDialog):
    def __init__(self, parent=None, quote=None, batches=None):
        super().__init__(parent)
        self.quote = quote
        self.batches = batches or []
        self.setWindowTitle("确认出库")
        self.setMinimumWidth(450)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        info_text = f"机型: {quote.get('series','')} {quote.get('cpu','')}\n客户: {quote.get('customer_name','')}\n数量: {quote.get('quote_quantity',1)} 台"
        self.info_label = QLabel(info_text)
        self.info_label.setObjectName("dialogInfoLabel")
        layout.addRow(self.info_label)

        self.batch_combo = QComboBox()
        for b in self.batches:
            self.batch_combo.addItem(
                f"批次#{b['id']} | 购入¥{b['purchase_price']:.0f} | 剩余{b['remaining']}台 | {b.get('date','')}",
                b["id"]
            )
        self.batch_combo.currentIndexChanged.connect(self._update_remaining)
        layout.addRow("扣减批次:", self.batch_combo)

        self.remaining_label = QLabel()
        self._update_remaining()
        layout.addRow("批次剩余:", self.remaining_label)

        self.sn_edit = QTextEdit()
        self.sn_edit.setPlaceholderText("条码枪扫码输入，每扫一条自动换行。也可手动输入，支持逗号/空格分隔。")
        self.sn_edit.setMaximumHeight(100)
        layout.addRow("出库SN:", self.sn_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("出库备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

        # 条码枪支持：拦截回车键，不让它提交对话框
        from PyQt6.QtCore import QEvent
        self.sn_edit.installEventFilter(self)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent, Qt
        if obj == self.sn_edit and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self.sn_edit.insertPlainText("\n")
                return True
        return super().eventFilter(obj, event)

    def _update_remaining(self):
        idx = self.batch_combo.currentIndex()
        if 0 <= idx < len(self.batches):
            remaining = self.batches[idx].get("remaining", 0)
            self.remaining_label.setText(f"{remaining} 台")
            color = "#D32F2F" if remaining < (self.quote.get("quote_quantity", 1) or 1) else "#388E3C"
            self.remaining_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def _validate_and_accept(self):
        quantity = self.quote.get("quote_quantity", 1) or 1
        idx = self.batch_combo.currentIndex()
        if idx < 0 or idx >= len(self.batches):
            QMessageBox.warning(self, "提示", "请选择批次")
            return
        remaining = self.batches[idx].get("remaining", 0)
        if remaining < quantity:
            QMessageBox.warning(self, "提示", f"批次剩余不足！需要 {quantity} 台，剩余 {remaining} 台")
            return
        self.accept()

    def get_data(self):
        idx = self.batch_combo.currentIndex()
        raw_sn = self.sn_edit.toPlainText().strip()
        sn_list = parse_sn_input(raw_sn)
        return {
            "batch_id": self.batch_combo.currentData() if idx >= 0 else None,
            "sn_list": ",".join(sn_list),
            "sn_count": len(sn_list),
            "remark": self.remark_edit.text().strip(),
        }



# ============================================================
# 收款/付款对话框
# ============================================================
class PaymentDialog(QDialog):
    def __init__(self, parent=None, title="收款", pay_type="receivable", quote=None, preview_pending=None):
        super().__init__(parent)
        self.pay_type = pay_type
        self.quote = quote
        self.preview_pending = preview_pending
        self.setWindowTitle(title)
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        if quote:
            total_amount = (quote.get("quote_price", 0) or 0) * (quote.get("quote_quantity", 1) or 1)
            received = quote.get("received_amount", 0) or 0
            remaining = total_amount - received
            info = f"客户: {quote.get('customer_name','')} | 总金额: ¥{total_amount:.0f} | 已收: ¥{received:.0f} | 待收: ¥{remaining:.0f}"
            info_label = QLabel(info)
            info_label.setObjectName("dialogInfoLabel")
            layout.addRow(info_label)

        if preview_pending:
            info_label = QLabel(f"当前待收金额: ¥{preview_pending:.0f}")
            info_label.setObjectName("dangerSummaryLabel")
            layout.addRow(info_label)

        self.amount_spin = QSpinBox()
        self.amount_spin.setRange(1, 999999)
        self.amount_spin.setPrefix("¥ ")
        self.amount_spin.setValue(0)
        layout.addRow("金额:", self.amount_spin)

        self.method_combo = QComboBox()
        self.method_combo.addItems(["微信", "支付宝", "转账", "现金"])
        layout.addRow("方式:", self.method_combo)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        layout.addRow("日期:", self.date_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "amount": self.amount_spin.value(),
            "method": self.method_combo.currentText(),
            "pay_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
        }


# ============================================================
# 收付款记录编辑对话框
# ============================================================
class PaymentEditDialog(QDialog):
    def __init__(self, parent=None, payment=None):
        super().__init__(parent)
        self.payment = payment
        self.setWindowTitle("编辑收付款记录")
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        # 显示原记录信息
        type_text = "收款" if payment.get("type") == "receivable" else "付款"
        obj_name = payment.get("customer_name", "") or payment.get("supplier_name", "")
        info = f"类型: {type_text} | 关联对象: {obj_name}"
        info_label = QLabel(info)
        info_label.setObjectName("dialogInfoLabel")
        layout.addRow(info_label)

        self.amount_spin = QSpinBox()
        self.amount_spin.setRange(1, 999999)
        self.amount_spin.setPrefix("¥ ")
        self.amount_spin.setValue(int(payment.get("amount", 0)))
        layout.addRow("金额:", self.amount_spin)

        self.method_combo = QComboBox()
        self.method_combo.addItems(["微信", "支付宝", "转账", "现金"])
        # 设置当前方式
        current_method = payment.get("method", "")
        idx = self.method_combo.findText(current_method)
        if idx >= 0:
            self.method_combo.setCurrentIndex(idx)
        layout.addRow("方式:", self.method_combo)

        self.date_edit = QDateEdit()
        if payment.get("pay_date"):
            qdate = QDate.fromString(payment.get("pay_date"), "yyyy-MM-dd")
            if qdate.isValid():
                self.date_edit.setDate(qdate)
        else:
            self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        layout.addRow("日期:", self.date_edit)

        self.remark_edit = QLineEdit()
        self.remark_edit.setText(payment.get("remark", "") or "")
        self.remark_edit.setPlaceholderText("备注（选填）")
        layout.addRow("备注:", self.remark_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "amount": self.amount_spin.value(),
            "method": self.method_combo.currentText(),
            "pay_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
        }


# ============================================================
# 客户对账单对话框
# ============================================================
class StatementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("客户对账单")
        self.setMinimumSize(800, 550)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("客户:"))

        self.customer_combo = QComboBox()
        self.customer_combo.setEditable(True)
        self.customer_combo.setPlaceholderText("选择或输入客户...")
        customers = get_all_customers()
        for c in customers:
            self.customer_combo.addItem(c["name"], c["id"])
        top_row.addWidget(self.customer_combo)

        top_row.addWidget(QLabel("开始:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        top_row.addWidget(self.date_from)

        top_row.addWidget(QLabel("结束:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        top_row.addWidget(self.date_to)

        self.query_btn = QPushButton("查询")
        self.query_btn.setObjectName("ghostBtn")
        self.query_btn.clicked.connect(self._query)
        top_row.addWidget(self.query_btn)

        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setObjectName("successBtn")
        self.export_btn.clicked.connect(self._export_excel)
        top_row.addWidget(self.export_btn)

        layout.addLayout(top_row)

        self.summary_label = QLabel()
        self.summary_label.setObjectName("summaryLabel")
        layout.addWidget(self.summary_label)

        self.statement_table = QTableWidget()
        self.statement_table.setAlternatingRowColors(True)
        self.statement_table.setColumnCount(11)
        self.statement_table.setHorizontalHeaderLabels(
            ["日期", "机型", "CPU", "数量", "购入价", "报价", "毛利", "状态", "已收款", "待收款", "备注"]
        )
        self.statement_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.statement_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.statement_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.statement_table)

    def _query(self):
        customer_name = self.customer_combo.currentText().strip()
        customer_id = self.customer_combo.currentData()
        if not customer_id and customer_name:
            matched = search_customers(customer_name)
            if matched:
                customer_id = matched[0]["id"]
            else:
                QMessageBox.warning(self, "提示", "未找到该客户")
                return
        if not customer_id:
            QMessageBox.warning(self, "提示", "请选择客户")
            return

        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        self._records = get_customer_statement(customer_id, date_from, date_to)

        self.statement_table.setRowCount(len(self._records))
        total_cost = 0
        total_sale = 0
        total_received = 0
        for i, r in enumerate(self._records):
            purchase_price = r.get("purchase_price", 0) or 0
            quote_price = r.get("quote_price", 0) or 0
            quantity = r.get("quote_quantity", 1) or 1
            received = r.get("received_amount", 0) or 0
            total_amount = quote_price * quantity
            profit = (quote_price - purchase_price) * quantity
            pending = total_amount - received

            self.statement_table.setItem(i, 0, QTableWidgetItem(r.get("quote_date", "")))
            self.statement_table.setItem(i, 1, QTableWidgetItem(r.get("series", "")))
            self.statement_table.setItem(i, 2, QTableWidgetItem(r.get("cpu", "")))
            self.statement_table.setItem(i, 3, QTableWidgetItem(str(quantity)))
            self.statement_table.setItem(i, 4, QTableWidgetItem(f"¥{purchase_price:.0f}"))
            self.statement_table.setItem(i, 5, QTableWidgetItem(f"¥{quote_price:.0f}"))
            profit_item = QTableWidgetItem(f"¥{profit:.0f}" if profit >= 0 else f"-¥{-profit:.0f}")
            profit_item.setForeground(QColor("#388E3C") if profit >= 0 else QColor("#D32F2F"))
            self.statement_table.setItem(i, 6, profit_item)
            self.statement_table.setItem(i, 7, QTableWidgetItem(r.get("status", "")))
            self.statement_table.setItem(i, 8, QTableWidgetItem(f"¥{received:.0f}"))
            self.statement_table.setItem(i, 9, QTableWidgetItem(f"¥{pending:.0f}" if pending > 0 else "已结清"))
            merged_remark = " | ".join(filter(None, [r.get("batch_remark", "") or "", r.get("remark", "") or ""]))
            self.statement_table.setItem(i, 10, QTableWidgetItem(merged_remark))

            total_cost += purchase_price * quantity
            total_sale += total_amount
            total_received += received

        self.statement_table.resizeColumnsToContents()

        total_pending = total_sale - total_received
        self.summary_label.setText(
            f"客户: {customer_name} | 总购入: ¥{total_cost:.0f} | 总金额: ¥{total_sale:.0f} | "
            f"已收款: ¥{total_received:.0f} | 待收款: ¥{total_pending:.0f} | 总毛利: ¥{total_sale - total_cost:.0f}"
        )

    def _export_excel(self):
        if not hasattr(self, '_records') or not self._records:
            QMessageBox.warning(self, "提示", "请先查询数据")
            return
        customer_name = self.customer_combo.currentText().strip()
        output = export_statement_to_excel(self._records, customer_name)
        QMessageBox.information(self, "导出成功", f"对账单已导出:\n{output}")


def export_statement_to_excel(records, customer_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(desktop, f"对账单_{customer_name}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"

    headers = ["日期", "机型", "CPU", "数量", "购入价", "报价", "毛利", "状态", "已收款", "待收款", "备注"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_alignment = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    total_cost = 0
    total_sale = 0
    total_received = 0
    for row_idx, r in enumerate(records, 2):
        purchase_price = r.get("purchase_price", 0) or 0
        quote_price = r.get("quote_price", 0) or 0
        quantity = r.get("quote_quantity", 1) or 1
        received = r.get("received_amount", 0) or 0
        total_amount = quote_price * quantity
        profit = (quote_price - purchase_price) * quantity
        pending = total_amount - received
        merged_remark = " | ".join(filter(None, [r.get("batch_remark", "") or "", r.get("remark", "") or ""]))

        row_data = [
            r.get("quote_date", ""), r.get("series", ""), r.get("cpu", ""),
            quantity, purchase_price, quote_price, profit, r.get("status", ""),
            received, pending if pending > 0 else "已结清", merged_remark,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        total_cost += purchase_price * quantity
        total_sale += total_amount
        total_received += received

    summary_row = len(records) + 2
    summary_data = [
        "合计", "", "", "",
        round(total_cost, 2), round(total_sale, 2),
        round(total_sale - total_cost, 2), "",
        round(total_received, 2), round(total_sale - total_received, 2), "",
    ]
    for col_idx, val in enumerate(summary_data, 1):
        cell = ws.cell(row=summary_row, column=col_idx, value=val)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border

    col_widths = [12, 16, 14, 8, 10, 10, 10, 10, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# 样式表
# ============================================================
APP_STYLE = """
/* ---- 全局基础 ---- */
QMainWindow { background-color: #F8FAFC; }
QWidget { font-family: 'Microsoft YaHei', 'Segoe UI', sans-serif; font-size: 13px; color: #1E293B; }

/* ---- 输入控件 ---- */
QLineEdit, QComboBox, QSpinBox, QDateEdit {
    padding: 7px 12px;
    border: 1px solid #E2E8F0;
    border-radius: 6px;
    background: #FFFFFF;
    color: #1E293B;
    font-size: 13px;
    selection-background-color: #DBEAFE;
}
QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDateEdit:focus { border-color: #3B82F6; }
QComboBox::drop-down { border: none; width: 24px; }
QComboBox::down-arrow {
    width: 10px; height: 10px; image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #64748B;
    margin-right: 6px;
}
QTextEdit {
    padding: 8px 12px; border: 1px solid #E2E8F0;
    border-radius: 6px; background: #FFFFFF;
    font-size: 13px; selection-background-color: #DBEAFE;
}

/* ---- 按钮基础 ---- */
QPushButton {
    padding: 6px 16px; border-radius: 6px;
    border: 1px solid #E2E8F0; background: #FFFFFF;
    color: #1E293B; font-size: 13px;
}
QPushButton:hover { background: #F1F5F9; border-color: #CBD5E1; }
QPushButton:pressed { background: #E2E8F0; }
QPushButton:disabled { color: #94A3B8; background: #F1F5F9; border-color: #E2E8F0; }

/* Primary */
QPushButton#primaryBtn { background: #3B82F6; color: #FFFFFF; border: 1px solid #3B82F6; }
QPushButton#primaryBtn:hover { background: #2563EB; border-color: #2563EB; }
QPushButton#primaryBtn:pressed { background: #1D4ED8; }

/* Success */
QPushButton#successBtn { background: #F0FDF4; color: #16A34A; border: 1px solid #BBF7D0; }
QPushButton#successBtn:hover { background: #DCFCE7; border-color: #86EFAC; }
QPushButton#successBtn:pressed { background: #BBF7D0; }

/* Warning */
QPushButton#warningBtn { background: #FFFBEB; color: #D97706; border: 1px solid #FDE68A; }
QPushButton#warningBtn:hover { background: #FEF3C7; border-color: #FCD34D; }
QPushButton#warningBtn:pressed { background: #FDE68A; }

/* Danger */
QPushButton#dangerBtn { background: #FEF2F2; color: #DC2626; border: 1px solid #FECACA; }
QPushButton#dangerBtn:hover { background: #FEE2E2; border-color: #FCA5A5; }
QPushButton#dangerBtn:pressed { background: #FECACA; }

/* Ghost */
QPushButton#ghostBtn { background: transparent; border: 1px solid #E2E8F0; color: #475569; }
QPushButton#ghostBtn:hover { background: #F8FAFC; border-color: #CBD5E1; color: #1E293B; }
QPushButton#ghostBtn:pressed { background: #F1F5F9; }

/* 表内操作按钮 */
QPushButton#tableActionPrimary {
    background: #EFF6FF; color: #2563EB; border: none;
    padding: 4px 12px; font-size: 12px; border-radius: 4px;
}
QPushButton#tableActionPrimary:hover { background: #DBEAFE; }
QPushButton#tableActionOrange {
    background: #FFFBEB; color: #D97706; border: none;
    padding: 4px 12px; font-size: 12px; border-radius: 4px;
}
QPushButton#tableActionOrange:hover { background: #FEF3C7; }
QPushButton#tableActionDanger {
    background: #FEF2F2; color: #DC2626; border: none;
    padding: 4px 12px; font-size: 12px; border-radius: 4px;
}
QPushButton#tableActionDanger:hover { background: #FEE2E2; }

/* 诊断选项按钮 */
QPushButton#diagnoseOptionBtn {
    text-align: left; padding: 8px 16px; font-size: 13px;
    border: 1px solid #E2E8F0; border-radius: 6px; background: #FFFFFF;
}
QPushButton#diagnoseOptionBtn:hover { background: #F1F5F9; border-color: #CBD5E1; }

/* ---- 表格 ---- */
QTableWidget {
    background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 8px;
    gridline-color: #F1F5F9; font-size: 13px;
    selection-background-color: #EFF6FF; selection-color: #1E293B;
    alternate-background-color: #F8FAFC;
}
QHeaderView::section {
    background: #F8FAFC; color: #64748B; font-weight: 600;
    font-size: 12px; padding: 8px 12px;
    border: none; border-bottom: 2px solid #E2E8F0;
}

/* ---- Tab ---- */
QTabWidget::pane {
    border: 1px solid #E2E8F0; border-radius: 0 0 8px 8px;
    background: #FFFFFF; padding: 0px;
}
QTabBar::tab {
    background: transparent; color: #64748B; padding: 10px 24px;
    font-size: 13px; font-weight: 500; border: none;
    border-bottom: 2px solid transparent; margin-right: 2px;
}
QTabBar::tab:selected { color: #1E293B; font-weight: 600; border-bottom: 2px solid #3B82F6; }
QTabBar::tab:hover:!selected { color: #1E293B; background: #F8FAFC; }

/* ---- GroupBox ---- */
QGroupBox {
    font-weight: 600; font-size: 13px; color: #1E293B;
    border: 1px solid #E2E8F0; border-radius: 8px;
    margin-top: 16px; padding: 20px 12px 12px 12px; background: #FFFFFF;
}
QGroupBox::title {
    subcontrol-origin: margin; left: 16px;
    padding: 0 8px; color: #1E293B;
}

/* ---- StatusBar ---- */
QStatusBar {
    background: #FFFFFF; border-top: 1px solid #E2E8F0;
    font-size: 12px; color: #64748B; padding: 2px 16px;
}

/* ---- Dialog ---- */
QDialog { background: #FFFFFF; }
QDialogButtonBox QPushButton { min-width: 80px; padding: 6px 24px; }

/* ---- Splitter ---- */
QSplitter::handle:horizontal { background: #E2E8F0; width: 1px; }
QSplitter::handle:vertical { background: #E2E8F0; height: 1px; }

/* ---- 语义化标签 ---- */
QLabel#sectionTitle { font-size: 15px; font-weight: 600; color: #1E293B; padding: 4px 0; }
QLabel#sectionTitleBlue { font-size: 15px; font-weight: 600; color: #3B82F6; padding: 4px 0; }
QLabel#sectionTitleRed { font-size: 15px; font-weight: 600; color: #DC2626; padding: 4px 0; }
QLabel#sectionTitleOrange { font-size: 15px; font-weight: 600; color: #D97706; padding: 4px 0; }
QLabel#summaryLabel { font-weight: 600; font-size: 14px; padding: 8px; color: #1E293B; }
QLabel#dangerSummaryLabel { font-weight: 600; font-size: 14px; padding: 8px; color: #DC2626; }
QLabel#dialogInfoLabel { font-size: 14px; font-weight: 600; padding: 8px; color: #1E293B; background: #F8FAFC; border-radius: 6px; }
QLabel#filterLabel { color: #64748B; font-size: 12px; }
QLabel#appTitle { font-size: 16px; font-weight: 700; color: #1E293B; padding: 0 4px; }
QTextEdit#reportText { font-size: 13px; font-family: 'Microsoft YaHei', monospace; padding: 12px; }

/* ---- 容器 ---- */
QFrame#filterCard { background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 8px; }
QFrame#toolbarSeparator { color: #E2E8F0; max-width: 1px; margin: 0 8px; }

/* ---- 搜索框 ---- */
QLineEdit#globalSearch {
    padding: 7px 12px 7px 32px; border: 1px solid #E2E8F0;
    border-radius: 6px; background: #F8FAFC; font-size: 13px;
}
QLineEdit#globalSearch:focus { background: #FFFFFF; border-color: #3B82F6; }
"""


# ============================================================
# 编辑机型对话框
# ============================================================
class ProductEditDialog(QDialog):
    def __init__(self, parent=None, product=None):
        super().__init__(parent)
        self.product = product  # dict or None
        self.setWindowTitle("编辑机型" if product else "新增机型")
        self.setMinimumWidth(450)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.series_edit = QLineEdit()
        self.cpu_edit = QLineEdit()
        self.ram_edit = QLineEdit()
        self.storage_edit = QLineEdit()
        self.gpu_edit = QLineEdit()
        self.screen_edit = QLineEdit()
        self.note_edit = QLineEdit()

        if product:
            self.series_edit.setText(product.get("series", ""))
            self.cpu_edit.setText(product.get("cpu", ""))
            self.ram_edit.setText(product.get("ram", ""))
            self.storage_edit.setText(product.get("storage", ""))
            self.gpu_edit.setText(product.get("gpu", ""))
            self.screen_edit.setText(product.get("screen", ""))
            self.note_edit.setText(product.get("note", ""))

        layout.addRow("系列:", self.series_edit)
        layout.addRow("CPU:", self.cpu_edit)
        layout.addRow("内存:", self.ram_edit)
        layout.addRow("硬盘:", self.storage_edit)
        layout.addRow("显卡:", self.gpu_edit)
        layout.addRow("屏幕:", self.screen_edit)
        layout.addRow("备注:", self.note_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "series": self.series_edit.text().strip(),
            "cpu": self.cpu_edit.text().strip(),
            "ram": self.ram_edit.text().strip(),
            "storage": self.storage_edit.text().strip(),
            "gpu": self.gpu_edit.text().strip(),
            "screen": self.screen_edit.text().strip(),
            "note": self.note_edit.text().strip(),
        }


# ============================================================
# 新增批次库存对话框
# ============================================================
class BatchDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新增库存批次")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.price_spin = QSpinBox()
        self.price_spin.setRange(0, 999999)
        self.price_spin.setPrefix("¥ ")
        self.price_spin.setValue(0)

        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 9999)
        self.quantity_spin.setValue(1)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setEditable(True)
        self.supplier_combo.setPlaceholderText("输入上游名称或选择...")
        self._load_suppliers()

        self.remark_edit = QLineEdit()
        self.remark_edit.setPlaceholderText("如: 含税/未税")

        self.sn_edit = QLineEdit()
        self.sn_edit.setPlaceholderText("序列号，多台用逗号分隔（选填）")

        layout.addRow("购入价:", self.price_spin)
        layout.addRow("数量:", self.quantity_spin)
        layout.addRow("入库日期:", self.date_edit)
        layout.addRow("上游:", self.supplier_combo)
        layout.addRow("备注:", self.remark_edit)
        layout.addRow("序列号:", self.sn_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _load_suppliers(self):
        suppliers = get_all_suppliers()
        for s in suppliers:
            self.supplier_combo.addItem(s["name"], s["id"])

    def get_data(self):
        supplier_name = self.supplier_combo.currentText().strip()
        supplier_id = self.supplier_combo.currentData()
        if not supplier_id and supplier_name:
            supplier_id = add_supplier(supplier_name)
        return {
            "price": self.price_spin.value(),
            "quantity": self.quantity_spin.value(),
            "date": self.date_edit.date().toString("yyyy-MM-dd"),
            "remark": self.remark_edit.text().strip(),
            "supplier_id": supplier_id,
            "sn_list": self.sn_edit.text().strip(),
        }


# ============================================================
# 客户编辑对话框
# ============================================================
class CustomerDialog(QDialog):
    def __init__(self, parent=None, customer=None):
        super().__init__(parent)
        self.customer = customer
        self.setWindowTitle("编辑客户" if customer else "新增客户")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.name_edit = QLineEdit()
        self.wechat_edit = QLineEdit()
        self.qq_edit = QLineEdit()
        self.phone_edit = QLineEdit()
        self.note_edit = QLineEdit()

        if customer:
            self.name_edit.setText(customer.get("name", ""))
            self.wechat_edit.setText(customer.get("wechat", ""))
            self.qq_edit.setText(customer.get("qq", ""))
            self.phone_edit.setText(customer.get("phone", ""))
            self.note_edit.setText(customer.get("note", ""))

        layout.addRow("名称:", self.name_edit)
        layout.addRow("微信:", self.wechat_edit)
        layout.addRow("QQ:", self.qq_edit)
        layout.addRow("电话:", self.phone_edit)
        layout.addRow("备注:", self.note_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_data(self):
        return {
            "name": self.name_edit.text().strip(),
            "wechat": self.wechat_edit.text().strip(),
            "qq": self.qq_edit.text().strip(),
            "phone": self.phone_edit.text().strip(),
            "note": self.note_edit.text().strip(),
        }


# ============================================================
# 报价记录编辑对话框
# ============================================================
class QuoteEditDialog(QDialog):
    def __init__(self, parent=None, quote=None, batch_id=None, customer_id=None):
        super().__init__(parent)
        self.quote = quote
        self.setWindowTitle("编辑报价记录" if quote else "新增报价记录")
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.price_spin = QSpinBox()
        self.price_spin.setRange(0, 999999)
        self.price_spin.setPrefix("¥ ")
        self.price_spin.setValue(0)

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.customer_combo = QComboBox()
        self.customer_combo.setEditable(True)
        self.customer_combo.setPlaceholderText("输入客户名称或选择...")
        self._load_customers()
        if customer_id:
            idx = self.customer_combo.findData(customer_id)
            if idx >= 0:
                self.customer_combo.setCurrentIndex(idx)

        self.remark_edit = QLineEdit()

        self.sn_edit = QLineEdit()
        self.sn_edit.setPlaceholderText("序列号，多台用逗号分隔（选填）")

        self.quantity_spin = QSpinBox()
        self.quantity_spin.setRange(1, 9999)
        self.quantity_spin.setValue(1)

        self.paid_combo = QComboBox()
        self.paid_combo.addItems(["否", "是"])

        if quote:
            self.price_spin.setValue(int(quote.get("quote_price", 0)))
            self.quantity_spin.setValue(quote.get("quote_quantity", 1))
            if quote.get("quote_date"):
                qdate = QDate.fromString(quote.get("quote_date"), "yyyy-MM-dd")
                if qdate.isValid():
                    self.date_edit.setDate(qdate)
            self.remark_edit.setText(quote.get("remark", ""))
            self.sn_edit.setText(quote.get("sn_list", "") or quote.get("batch_sn_list", ""))
            paid_value = quote.get("paid", "否")
            idx = self.paid_combo.findText(paid_value)
            if idx >= 0:
                self.paid_combo.setCurrentIndex(idx)

        layout.addRow("对外报价:", self.price_spin)
        layout.addRow("数量:", self.quantity_spin)
        layout.addRow("报价日期:", self.date_edit)
        layout.addRow("客户:", self.customer_combo)
        layout.addRow("备注:", self.remark_edit)
        layout.addRow("序列号:", self.sn_edit)
        layout.addRow("是否打款:", self.paid_combo)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _load_customers(self):
        customers = get_all_customers()
        for c in customers:
            self.customer_combo.addItem(c["name"], c["id"])

    def get_data(self):
        customer_name = self.customer_combo.currentText().strip()
        customer_id = self.customer_combo.currentData()
        if not customer_id and customer_name:
            customer_id = add_customer(customer_name)
        return {
            "quote_price": self.price_spin.value(),
            "quote_quantity": self.quantity_spin.value(),
            "quote_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "customer_id": customer_id,
            "remark": self.remark_edit.text().strip(),
            "sn_list": self.sn_edit.text().strip(),
            "paid": self.paid_combo.currentText(),
            "batch_id": self.quote.get("batch_id") if self.quote else None,
        }


# ============================================================
# 报价面板
# ============================================================
class QuotePanel(QWidget):
    """报价操作面板，显示选中机型的批次详情并进行报价"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_product_id = None
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # 机型信息
        self.product_label = QLabel("未选择机型")
        self.product_label.setObjectName("sectionTitleBlue")
        layout.addWidget(self.product_label)

        # 批次表格
        self.batch_table = QTableWidget()
        self.batch_table.setAlternatingRowColors(True)
        self.batch_table.setColumnCount(7)
        self.batch_table.setHorizontalHeaderLabels(["购入价", "进货数量", "剩余", "入库日期", "上游", "备注", "序列号"])
        self.batch_table.horizontalHeader().setStretchLastSection(True)
        self.batch_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.batch_table.setMinimumHeight(120)
        layout.addWidget(QLabel("库存批次:"))
        layout.addWidget(self.batch_table)

        # 批次操作按钮行
        btn_row1 = QHBoxLayout()
        self.add_batch_btn = QPushButton("+ 新增批次")
        self.add_batch_btn.setObjectName("primaryBtn")
        self.add_batch_btn.clicked.connect(self.on_add_batch)
        self.del_batch_btn = QPushButton("删除批次")
        self.del_batch_btn.setObjectName("dangerBtn")
        self.del_batch_btn.clicked.connect(self.on_delete_batch)
        self.refresh_batch_btn = QPushButton("刷新")
        self.refresh_batch_btn.setObjectName("ghostBtn")
        self.refresh_batch_btn.clicked.connect(self.refresh)
        btn_row1.addWidget(self.add_batch_btn)
        btn_row1.addWidget(self.del_batch_btn)
        btn_row1.addWidget(self.refresh_batch_btn)
        btn_row1.addStretch()
        layout.addLayout(btn_row1)

        # 报价操作区
        layout.addSpacing(10)
        group = QGroupBox("报价操作")
        glayout = QGridLayout(group)

        self.quote_price_spin = QSpinBox()
        self.quote_price_spin.setRange(0, 999999)
        self.quote_price_spin.setPrefix("¥ ")
        self.quote_price_spin.setValue(0)

        self.quote_quantity_spin = QSpinBox()
        self.quote_quantity_spin.setRange(1, 9999)
        self.quote_quantity_spin.setValue(1)

        self.customer_combo = QComboBox()
        self.customer_combo.setEditable(True)
        self.customer_combo.setPlaceholderText("输入客户名称或选择...")
        self.customer_combo.setMinimumWidth(150)

        self.quote_remark_edit = QLineEdit()
        self.quote_remark_edit.setPlaceholderText("备注（选填）")

        self.quote_paid_combo = QComboBox()
        self.quote_paid_combo.addItems(["否", "是"])

        glayout.addWidget(QLabel("对外报价:"), 0, 0)
        glayout.addWidget(self.quote_price_spin, 0, 1)
        glayout.addWidget(QLabel("报价数量:"), 0, 2)
        glayout.addWidget(self.quote_quantity_spin, 0, 3)
        glayout.addWidget(QLabel("客户:"), 1, 0)
        glayout.addWidget(self.customer_combo, 1, 1)
        glayout.addWidget(QLabel("是否打款:"), 1, 2)
        glayout.addWidget(self.quote_paid_combo, 1, 3)
        glayout.addWidget(QLabel("备注:"), 2, 0)
        glayout.addWidget(self.quote_remark_edit, 2, 1, 1, 3)

        btn_row2 = QHBoxLayout()
        self.quote_copy_btn = QPushButton("报价并复制文本")
        self.quote_copy_btn.setObjectName("primaryBtn")
        self.quote_copy_btn.clicked.connect(self.on_quote_and_copy)
        self.quote_img_btn = QPushButton("生成报价图片")
        self.quote_img_btn.setObjectName("ghostBtn")
        self.quote_img_btn.clicked.connect(self.on_quote_image)
        btn_row2.addWidget(self.quote_copy_btn)
        btn_row2.addWidget(self.quote_img_btn)
        btn_row2.addStretch()
        glayout.addLayout(btn_row2, 3, 0, 1, 4)

        layout.addWidget(group)
        layout.addStretch()

    def load_product(self, product_id, series, cpu, ram, storage, gpu, screen, note):
        self.current_product_id = product_id
        self.product_specs = {
            "series": series, "cpu": cpu, "ram": ram,
            "storage": storage, "gpu": gpu, "screen": screen, "note": note,
        }
        parts = [series]
        if cpu: parts.append(cpu)
        if ram: parts.append(ram)
        if storage: parts.append(storage)
        if gpu: parts.append(gpu)
        self.product_label.setText("  ".join(parts))
        self.refresh()

        # Skill 1: 加载报价建议
        try:
            from src.utils.quote_assist import suggest_price
            suggestion = suggest_price(series=series, cpu=cpu, ram=ram, storage=storage, gpu=gpu)
            if suggestion and suggestion.get("suggested_mid", 0) > 0:
                self.quote_price_spin.setValue(suggestion["suggested_mid"])
        except Exception:
            pass

    def refresh(self):
        if not self.current_product_id:
            return
        
        batches = get_batches(self.current_product_id)
        
        suppliers = {s["id"]: s["name"] for s in get_all_suppliers()}
        
        self.batch_table.setRowCount(len(batches))
        for i, b in enumerate(batches):
            price_item = QTableWidgetItem(f"¥ {b['purchase_price']:.0f}")
            price_item.setData(Qt.ItemDataRole.UserRole, b["id"])
            self.batch_table.setItem(i, 0, price_item)
            self.batch_table.setItem(i, 1, QTableWidgetItem(str(b['quantity'])))
            self.batch_table.setItem(i, 2, QTableWidgetItem(str(b['remaining'])))
            self.batch_table.setItem(i, 3, QTableWidgetItem(b['date']))
            supplier_name = suppliers.get(b.get('supplier_id'), '')
            self.batch_table.setItem(i, 4, QTableWidgetItem(supplier_name))
            self.batch_table.setItem(i, 5, QTableWidgetItem(b.get('remark', '')))
            self.batch_table.setItem(i, 6, QTableWidgetItem(b.get('sn_list', '')))
        self.batch_table.resizeColumnsToContents()

        self.customer_combo.clear()
        customers = get_all_customers()
        for c in customers:
            self.customer_combo.addItem(c["name"], c["id"])

    def on_add_batch(self):
        if not self.current_product_id:
            QMessageBox.warning(self, "提示", "请先选择机型")
            return
        dlg = BatchDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            add_batch(
                self.current_product_id,
                data["price"],
                data["quantity"],
                data["quantity"],
                data["date"],
                data["remark"],
                data["supplier_id"],
                data["sn_list"],
            )
            add_operation_log("入库", "batches", 0, f"数量={data['quantity']}, 单价={data['price']}")
            self.refresh()

    def on_delete_batch(self):
        row = self.batch_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择要删除的批次")
            return
        price_item = self.batch_table.item(row, 0)
        if not price_item:
            return
        batch_id = price_item.data(Qt.ItemDataRole.UserRole)
        if batch_id is None:
            return
        reply = QMessageBox.question(
            self, "确认删除", "确定删除此批次？此操作不可恢复。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            delete_batch(batch_id)
            add_operation_log("删除批次", "batches", batch_id, "删除批次")
            self.refresh()

    def on_quote_and_copy(self):
        """报价并复制文本到剪贴板"""
        if not self._validate_quote():
            return
        product, batch, quote_price, quote_quantity, customer_name, remark = self._prepare_quote_data()
        success, message = self._record_quote(batch["id"], quote_price, quote_quantity, remark)
        
        if not success:
            QMessageBox.warning(self, "报价失败", message)
            return
        
        text = self._format_quote_text(product, quote_price, customer_name, quote_quantity)
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(self, "成功", f"{message}\n报价已复制到剪贴板！")
        self.refresh()

    def on_quote_image(self):
        """生成报价图片"""
        if not self._validate_quote():
            return
        product, batch, quote_price, quote_quantity, customer_name, remark = self._prepare_quote_data()
        success, message = self._record_quote(batch["id"], quote_price, quote_quantity, remark)
        
        if not success:
            QMessageBox.warning(self, "报价失败", message)
            return
        
        total = quote_price * quote_quantity
        output = generate_single_quote_card(
            series=product.get("series", ""),
            cpu=product.get("cpu", ""),
            ram=product.get("ram", ""),
            storage=product.get("storage", ""),
            gpu=product.get("gpu", ""),
            screen=product.get("screen", ""),
            note=product.get("note", ""),
            customer_name=customer_name,
            quote_price=f"¥ {quote_price:.0f}",
            quote_quantity=quote_quantity,
            total_price=f"¥ {total:.0f}",
        )
        QMessageBox.information(self, "成功", f"{message}\n报价图片已生成:\n{output}")
        self.refresh()

    def _validate_quote(self):
        if not self.current_product_id:
            QMessageBox.warning(self, "提示", "请先在左侧选择机型")
            return False
        if self.quote_price_spin.value() <= 0:
            QMessageBox.warning(self, "提示", "请输入对外报价")
            return False
        row = self.batch_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请选择一个库存批次")
            return False
        
        # 从表格行的 UserRole 中获取 batch_id
        price_item = self.batch_table.item(row, 0)
        if not price_item:
            return False
        batch_id = price_item.data(Qt.ItemDataRole.UserRole)
        if batch_id is None:
            return False
        
        from src.models.database import get_batch_remaining
        quote_quantity = self.quote_quantity_spin.value()
        remaining = get_batch_remaining(batch_id)
        
        if remaining < quote_quantity:
            QMessageBox.warning(
                self, "库存不足", 
                f"库存不足！\n\n当前批次剩余: {remaining} 台\n本次报价数量: {quote_quantity} 台"
            )
            return False
        
        return True

    def _prepare_quote_data(self):
        row = self.batch_table.currentRow()
        if row < 0:
            return None, None, None, None, None, None
        # 从表格行的 UserRole 中获取 batch_id，避免重新查询导致索引不一致
        price_item = self.batch_table.item(row, 0)
        if not price_item:
            return None, None, None, None, None, None
        batch_id = price_item.data(Qt.ItemDataRole.UserRole)
        if batch_id is None:
            return None, None, None, None, None, None
        from src.models.database import get_connection
        conn = get_connection()
        batch_row = conn.execute(
            "SELECT id, product_id, purchase_price, quantity, remaining, date, remark, supplier_id, sn_list FROM batches WHERE id=?",
            (batch_id,),
        ).fetchone()
        conn.close()
        if not batch_row:
            return None, None, None, None, None, None
        batch = dict(batch_row)
        product = getattr(self, 'product_specs', {})
        quote_price = self.quote_price_spin.value()
        quote_quantity = self.quote_quantity_spin.value()
        customer_name = self.customer_combo.currentText().strip()
        remark = self.quote_remark_edit.text().strip()
        return product, batch, quote_price, quote_quantity, customer_name, remark

    def _record_quote(self, batch_id, quote_price, quote_quantity, remark):
        customer_name = self.customer_combo.currentText().strip()
        customer_id = None
        if customer_name:
            customers = search_customers(customer_name)
            matched = [c for c in customers if c["name"] == customer_name]
            if matched:
                customer_id = matched[0]["id"]
            else:
                customer_id = add_customer(customer_name)

        today = datetime.now().strftime("%Y-%m-%d")
        paid = self.quote_paid_combo.currentText()
        add_quote(batch_id, customer_id, quote_price, quote_quantity, today, remark, paid)
        
        return True, "报价已保存（待确认）"

    def _format_quote_text(self, product, quote_price, customer_name, quote_quantity=1):
        parts = [product.get("series", "")]
        if product.get("cpu"): parts.append(product["cpu"])
        if product.get("ram"): parts.append(product["ram"])
        if product.get("storage"): parts.append(product["storage"])
        if product.get("gpu"): parts.append(product["gpu"])
        if product.get("screen"): parts.append(product["screen"])
        if product.get("note"): parts.append(product["note"])

        spec = " ".join(parts)
        if quote_quantity > 1:
            total = quote_price * quote_quantity
            price_text = f"报价: ¥{quote_price:.0f} × {quote_quantity} = ¥{total:.0f}"
        else:
            price_text = f"报价: ¥{quote_price:.0f}"
        text = f"{spec}\n{price_text}"
        if customer_name:
            text = f"{customer_name}\n{text}"
        text += f"\n{WATERMARK_TEXT}"
        return text


# ============================================================
# 操作日志对话框
# ============================================================
class OperationLogDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("最近操作记录")
        self.setMinimumSize(700, 400)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["时间", "操作", "对象", "描述"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSortingEnabled(True)

        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        refresh_btn = QPushButton("刷新")
        refresh_btn.setObjectName("ghostBtn")
        refresh_btn.clicked.connect(self.load_logs)
        btn_row.addWidget(refresh_btn)
        close_btn = QPushButton("关闭")
        close_btn.setObjectName("ghostBtn")
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self.load_logs()

    def load_logs(self):
        from src.models.database import get_operation_logs
        logs = get_operation_logs(limit=50)
        self.table.setRowCount(len(logs))
        for i, log in enumerate(logs):
            self.table.setItem(i, 0, QTableWidgetItem(log.get("created_at", "")))
            self.table.setItem(i, 1, QTableWidgetItem(log.get("operation", "")))
            self.table.setItem(i, 2, QTableWidgetItem(
                f"{log.get('table_name', '')}(ID:{log.get('record_id', '')})"
            ))
            self.table.setItem(i, 3, QTableWidgetItem(log.get("description", "")))
        self.table.resizeColumnsToContents()


# ============================================================
# 主窗口
# ============================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("调货助手 v1.10")
        self.setMinimumSize(1100, 700)

        # 初始化数据库（含自动备份 + 完整性检查）
        init_ok, init_msg = init_db()
        self.backup_status = init_msg

        # 中心控件
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(12, 12, 12, 12)

        # 搜索 + 工具栏
        top_bar = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setObjectName("globalSearch")
        self.search_edit.setPlaceholderText("搜索机型（系列/CPU/关键字）...")
        self.search_edit.textChanged.connect(self.on_search)
        self.search_edit.setMinimumWidth(250)

        self.import_btn = QPushButton("导入 Word 价格表")
        self.import_btn.setObjectName("ghostBtn")
        self.import_btn.clicked.connect(self.on_import_word)
        self.broadcast_btn = QPushButton("群发图片")
        self.broadcast_btn.setObjectName("ghostBtn")
        self.broadcast_btn.clicked.connect(self.on_broadcast)
        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setObjectName("ghostBtn")
        self.export_btn.clicked.connect(self.on_export_excel)
        self.export_json_btn = QPushButton("JSON备份")
        self.export_json_btn.setObjectName("ghostBtn")
        self.export_json_btn.clicked.connect(self.on_export_json)
        self.import_json_btn = QPushButton("JSON导入")
        self.import_json_btn.setObjectName("ghostBtn")
        self.import_json_btn.clicked.connect(self.on_import_json)
        self.log_btn = QPushButton("操作日志")
        self.log_btn.setObjectName("ghostBtn")
        self.log_btn.clicked.connect(self.on_show_logs)
        self.statement_btn = QPushButton("对账单")
        self.statement_btn.setObjectName("ghostBtn")
        self.statement_btn.clicked.connect(self.on_statement)
        self.follow_up_btn = QPushButton("🔔 跟单提醒")
        self.follow_up_btn.setObjectName("ghostBtn")
        self.follow_up_btn.clicked.connect(self.on_follow_up)
        self.report_btn = QPushButton("📊 月度报告")
        self.report_btn.setObjectName("ghostBtn")
        self.report_btn.clicked.connect(self.on_monthly_report)
        self.diagnose_btn = QPushButton("🔧 远程诊断")
        self.diagnose_btn.setObjectName("ghostBtn")
        self.diagnose_btn.clicked.connect(self.on_remote_diagnose)
        self.price_diff_btn = QPushButton("价格异动")
        self.price_diff_btn.setObjectName("ghostBtn")
        self.price_diff_btn.clicked.connect(self.on_price_diff)
        self.quote_assist_btn = QPushButton("报价助手")
        self.quote_assist_btn.setObjectName("ghostBtn")
        self.quote_assist_btn.clicked.connect(self.on_quote_assist)
        self.shipment_flow_btn = QPushButton("出库一条龙")
        self.shipment_flow_btn.setObjectName("ghostBtn")
        self.shipment_flow_btn.clicked.connect(self.on_shipment_flow)

        top_bar.addWidget(QLabel("🔍"))
        top_bar.addWidget(self.search_edit)
        top_bar.addStretch()

        def _add_sep(layout):
            sep = QFrame()
            sep.setFrameShape(QFrame.Shape.VLine)
            sep.setObjectName("toolbarSeparator")
            layout.addWidget(sep)

        top_bar.addWidget(self.follow_up_btn)
        top_bar.addWidget(self.report_btn)
        top_bar.addWidget(self.diagnose_btn)
        top_bar.addWidget(self.price_diff_btn)
        top_bar.addWidget(self.quote_assist_btn)
        top_bar.addWidget(self.shipment_flow_btn)
        _add_sep(top_bar)
        top_bar.addWidget(self.import_btn)
        top_bar.addWidget(self.broadcast_btn)
        top_bar.addWidget(self.export_btn)
        top_bar.addWidget(self.statement_btn)
        top_bar.addWidget(self.export_json_btn)
        top_bar.addWidget(self.import_json_btn)
        _add_sep(top_bar)
        top_bar.addWidget(self.log_btn)
        main_layout.addLayout(top_bar)

        # Tab 页面
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # === Tab 1: 机型管理 ===
        tab_products = QWidget()
        tab_layout = QHBoxLayout(tab_products)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 左：机型列表
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        btn_row = QHBoxLayout()
        self.add_product_btn = QPushButton("+ 新增机型")
        self.add_product_btn.setObjectName("primaryBtn")
        self.add_product_btn.clicked.connect(self.on_add_product)
        self.edit_product_btn = QPushButton("编辑")
        self.edit_product_btn.setObjectName("ghostBtn")
        self.edit_product_btn.clicked.connect(self.on_edit_product)
        self.del_product_btn = QPushButton("删除")
        self.del_product_btn.setObjectName("dangerBtn")
        self.del_product_btn.clicked.connect(self.on_delete_product)
        btn_row.addWidget(self.add_product_btn)
        btn_row.addWidget(self.edit_product_btn)
        btn_row.addWidget(self.del_product_btn)
        self.refresh_product_list_btn = QPushButton("刷新")
        self.refresh_product_list_btn.setObjectName("ghostBtn")
        self.refresh_product_list_btn.clicked.connect(lambda: self.refresh_product_list(self.search_edit.text().strip()))
        btn_row.addWidget(self.refresh_product_list_btn)
        btn_row.addStretch()
        left_layout.addLayout(btn_row)

        self.product_table = QTableWidget()
        self.product_table.setAlternatingRowColors(True)
        self.product_table.setColumnCount(9)
        self.product_table.setHorizontalHeaderLabels(["ID", "系列", "CPU", "内存", "硬盘", "显卡", "屏幕", "备注", "库存"])
        self.product_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.product_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.product_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.product_table.horizontalHeader().setStretchLastSection(True)
        self.product_table.setColumnHidden(0, True)  # 隐藏 ID 列
        self.product_table.itemSelectionChanged.connect(self.on_product_selected)
        left_layout.addWidget(self.product_table)

        splitter.addWidget(left_widget)

        # 右：报价面板
        self.quote_panel = QuotePanel()
        splitter.addWidget(self.quote_panel)
        splitter.setSizes([600, 450])

        tab_layout.addWidget(splitter)
        self.tabs.addTab(tab_products, "📋 机型管理")

        # === Tab 2: 客户管理 ===
        tab_customers = self._build_customer_tab()
        self.tabs.addTab(tab_customers, "👤 客户管理")

        # === Tab 3: 上游管理 ===
        tab_suppliers = self._build_supplier_tab()
        self.tabs.addTab(tab_suppliers, "🏭 上游管理")

        # === Tab 4: 报价记录 ===
        tab_records = self._build_records_tab()
        self.tabs.addTab(tab_records, "📊 报价记录")

        # === Tab 5: 账款管理 ===
        tab_finance = self._build_finance_tab()
        self.tabs.addTab(tab_finance, "💰 账款管理")

        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_label = QLabel(f"就绪 | {self.backup_status}")
        self.status_bar.addWidget(self.status_label)

        # 加载数据
        self.refresh_product_list()
        self.refresh_customer_list()
        self.refresh_supplier_list()
        self.refresh_records()

        # 启用所有表格的列头排序
        for table in self.findChildren(QTableWidget):
            table.setSortingEnabled(True)

    def on_confirm_quote(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        quote = get_quote_by_id(quote_id)
        if not quote:
            return
        status = quote.get("status", "")
        if status != "待确认":
            QMessageBox.warning(self, "提示", f"当前状态为「{status}」，只有「待确认」状态的订单可以确认报价")
            return
        reply = QMessageBox.question(
            self, "确认报价", f"确定将此订单状态更新为「已报价」？\n\n客户: {quote.get('customer_name','')}\n机型: {quote.get('series','')} {quote.get('cpu','')}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            update_quote_status(quote_id, "已报价")
            self.refresh_records()

    def on_ship_quote(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        quote = get_quote_by_id(quote_id)
        if not quote:
            return
        status = quote.get("status", "")
        if status not in ("待确认", "已报价"):
            QMessageBox.warning(self, "提示", f"当前状态为「{status}」，只有「待确认」或「已报价」状态的订单可以出库")
            return

        from src.models.database import get_connection
        conn = get_connection()
        batch_row = conn.execute("SELECT product_id FROM batches WHERE id=?", (quote.get("batch_id"),)).fetchone()
        product_id = batch_row[0] if batch_row else None
        conn.close()

        batches = get_batches(product_id) if product_id else []
        if not batches:
            QMessageBox.warning(self, "提示", "没有可用批次，无法出库")
            return

        dlg = ShipmentDialog(self, quote, batches)
        if dlg.exec():
            data = dlg.get_data()
            # 使用封装好的 ship_quote 函数（含校验+扣减+保存SN+状态更新）
            success, msg = ship_quote(quote_id, data.get("sn_list", ""))
            if success:
                QMessageBox.information(self, "成功", msg)
                add_operation_log("出库", "quotes", quote_id, f"SN={data.get('sn_list', '')}")
                self.refresh_records()
            else:
                QMessageBox.warning(self, "出库失败", msg)

    def on_receive_payment(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        quote = get_quote_by_id(quote_id)
        if not quote:
            return
        status = quote.get("status", "")
        if status in ("已取消",):
            QMessageBox.warning(self, "提示", f"当前状态为「{status}」，无法收款")
            return

        dlg = PaymentDialog(self, title="收款", pay_type="receivable", quote=quote)
        if dlg.exec():
            data = dlg.get_data()
            if data["amount"] <= 0:
                QMessageBox.warning(self, "提示", "请输入收款金额")
                return
            add_payment(
                quote_id=quote_id,
                customer_id=quote.get("customer_id"),
                pay_type="receivable",
                amount=data["amount"],
                pay_date=data["pay_date"],
                method=data["method"],
                remark=data["remark"],
            )
            QMessageBox.information(self, "成功", f"收款 ¥{data['amount']:.2f} 已记录！")
            add_operation_log("收款", "quotes", quote_id, f"金额={data['amount']:.2f}")
            self.refresh_records()

    def on_cancel_quote(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        quote = get_quote_by_id(quote_id)
        if not quote:
            return
        status = quote.get("status", "")
        if status == "已收款":
            QMessageBox.warning(self, "提示", "已收款的订单不能取消，如需退款请手动处理")
            return
        reply = QMessageBox.question(
            self, "确认取消",
            f"确定取消此订单？\n\n客户: {quote.get('customer_name','')}\n机型: {quote.get('series','')} {quote.get('cpu','')}\n当前状态: {status}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            update_quote_status(quote_id, "已取消")
            add_operation_log("取消订单", "quotes", quote_id, "取消订单")
            self.refresh_records()

    def on_statement(self):
        dlg = StatementDialog(self)
        dlg.exec()

    # -------------------------------------------------------
    # 账款管理
    # -------------------------------------------------------
    def _build_finance_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)

        # === 上部分：应收/应付（占比调大） ===
        top_splitter = QSplitter(Qt.Orientation.Horizontal)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        left_title = QLabel("应收款项（客户欠款）")
        left_title.setObjectName("sectionTitleRed")
        left_layout.addWidget(left_title)

        self.receivable_table = QTableWidget()
        self.receivable_table.setAlternatingRowColors(True)
        self.receivable_table.setColumnCount(5)
        self.receivable_table.setHorizontalHeaderLabels(["客户", "欠款金额", "订单数", "联系方式", "操作"])
        self.receivable_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.receivable_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.receivable_table.horizontalHeader().setStretchLastSection(True)
        self.receivable_table.setMinimumHeight(200)  # 设置最小高度
        left_layout.addWidget(self.receivable_table)

        self.refresh_receivable_btn = QPushButton("刷新")
        self.refresh_receivable_btn.setObjectName("ghostBtn")
        self.refresh_receivable_btn.clicked.connect(self.refresh_finance)
        left_btn_row = QHBoxLayout()
        left_btn_row.addStretch()
        left_btn_row.addWidget(self.refresh_receivable_btn)
        left_layout.addLayout(left_btn_row)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        right_title = QLabel("应付款项（欠上游款）")
        right_title.setObjectName("sectionTitleOrange")
        right_layout.addWidget(right_title)

        self.payable_table = QTableWidget()
        self.payable_table.setAlternatingRowColors(True)
        self.payable_table.setColumnCount(5)
        self.payable_table.setHorizontalHeaderLabels(["上游", "欠款金额", "批次", "联系方式", "操作"])
        self.payable_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.payable_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.payable_table.horizontalHeader().setStretchLastSection(True)
        self.payable_table.setMinimumHeight(200)  # 设置最小高度
        right_layout.addWidget(self.payable_table)

        self.refresh_payable_btn = QPushButton("刷新")
        self.refresh_payable_btn.setObjectName("ghostBtn")
        self.refresh_payable_btn.clicked.connect(self.refresh_finance)
        right_btn_row = QHBoxLayout()
        right_btn_row.addStretch()
        right_btn_row.addWidget(self.refresh_payable_btn)
        right_layout.addLayout(right_btn_row)

        top_splitter.addWidget(left_widget)
        top_splitter.addWidget(right_widget)
        top_splitter.setSizes([400, 400])
        # 上部分占比 3（75%）
        layout.addWidget(top_splitter, stretch=3)

        # === 下部分：收付款流水预览（占比调小） ===
        flow_group = QGroupBox("收付款流水记录")
        flow_group.setMaximumHeight(250)  # 设置最大高度，限制流水区域
        flow_layout = QVBoxLayout(flow_group)

        # 筛选行
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("类型:"))
        self.payment_type_filter = QComboBox()
        self.payment_type_filter.addItems(["全部", "收款", "付款"])
        self.payment_type_filter.currentTextChanged.connect(self.refresh_payment_flow)
        filter_row.addWidget(self.payment_type_filter)

        filter_row.addWidget(QLabel("对象:"))
        self.payment_object_filter = QComboBox()
        self.payment_object_filter.setEditable(True)
        self.payment_object_filter.setPlaceholderText("全部对象...")
        self.payment_object_filter.currentTextChanged.connect(self.refresh_payment_flow)
        filter_row.addWidget(self.payment_object_filter)

        filter_row.addWidget(QLabel("日期从:"))
        self.payment_date_from = QDateEdit()
        self.payment_date_from.setCalendarPopup(True)
        self.payment_date_from.setDate(QDate.currentDate().addMonths(-3))
        self.payment_date_from.dateChanged.connect(self.refresh_payment_flow)
        filter_row.addWidget(self.payment_date_from)

        filter_row.addWidget(QLabel("至:"))
        self.payment_date_to = QDateEdit()
        self.payment_date_to.setCalendarPopup(True)
        self.payment_date_to.setDate(QDate.currentDate())
        self.payment_date_to.dateChanged.connect(self.refresh_payment_flow)
        filter_row.addWidget(self.payment_date_to)

        self.refresh_flow_btn = QPushButton("刷新")
        self.refresh_flow_btn.setObjectName("ghostBtn")
        self.refresh_flow_btn.clicked.connect(self.refresh_payment_flow)
        filter_row.addWidget(self.refresh_flow_btn)

        filter_row.addStretch()
        flow_layout.addLayout(filter_row)

        # 流水表格
        self.payment_flow_table = QTableWidget()
        self.payment_flow_table.setAlternatingRowColors(True)
        self.payment_flow_table.setColumnCount(8)
        self.payment_flow_table.setHorizontalHeaderLabels(
            ["ID", "日期", "类型", "金额", "方式", "关联对象", "备注", "操作"]
        )
        self.payment_flow_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.payment_flow_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.payment_flow_table.horizontalHeader().setStretchLastSection(True)
        self.payment_flow_table.setColumnHidden(0, True)  # 隐藏 ID 列
        flow_layout.addWidget(self.payment_flow_table)

        # 统计标签
        self.payment_flow_stats = QLabel()
        self.payment_flow_stats.setObjectName("summaryLabel")
        flow_layout.addWidget(self.payment_flow_stats)

        # 下部分占比 1（25%）
        layout.addWidget(flow_group, stretch=1)
        return tab

    def refresh_finance(self):
        """刷新应收/应付表格"""
        from src.models.database import get_connection
        conn = get_connection()

        rows = conn.execute("""
            SELECT c.id, c.name, c.wechat, c.phone,
                   COALESCE(SUM(q.quote_price * q.quote_quantity - q.received_amount), 0) as debt,
                   COUNT(q.id) as order_count
            FROM customers c
            LEFT JOIN quotes q ON q.customer_id = c.id AND q.status IN ('已报价','已出库')
            GROUP BY c.id
            HAVING debt > 0
            ORDER BY debt DESC
        """).fetchall()

        self.receivable_table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            self.receivable_table.setItem(i, 0, QTableWidgetItem(r[1]))
            debt_item = QTableWidgetItem(f"¥{r[4]:.0f}")
            debt_item.setForeground(QColor("#D32F2F"))
            self.receivable_table.setItem(i, 1, debt_item)
            self.receivable_table.setItem(i, 2, QTableWidgetItem(str(r[5])))
            contact = " | ".join(filter(None, [r[2] or "", r[3] or ""]))
            self.receivable_table.setItem(i, 3, QTableWidgetItem(contact))
            pay_btn = QPushButton("收款")
            pay_btn.setObjectName("tableActionPrimary")
            pay_btn.clicked.connect(lambda checked, cid=r[0]: self._on_finance_receive(cid))
            self.receivable_table.setCellWidget(i, 4, pay_btn)
        self.receivable_table.resizeColumnsToContents()

        s_rows = conn.execute("""
            SELECT s.id, s.name, s.wechat, s.phone,
                   COALESCE(SUM(b.purchase_price * b.quantity), 0) -
                   COALESCE((SELECT SUM(py.amount) FROM payments py WHERE py.supplier_id = s.id AND py.type = 'payable'), 0) as debt
            FROM suppliers s
            LEFT JOIN batches b ON b.supplier_id = s.id
            GROUP BY s.id
            HAVING debt > 0
            ORDER BY debt DESC
        """).fetchall()

        self.payable_table.setRowCount(len(s_rows))
        for i, r in enumerate(s_rows):
            self.payable_table.setItem(i, 0, QTableWidgetItem(r[1]))
            debt_item = QTableWidgetItem(f"¥{r[4]:.0f}")
            debt_item.setForeground(QColor("#F57C00"))
            self.payable_table.setItem(i, 1, debt_item)
            batch_count = conn.execute(
                "SELECT COUNT(*) FROM batches WHERE supplier_id=?", (r[0],)
            ).fetchone()[0]
            self.payable_table.setItem(i, 2, QTableWidgetItem(str(batch_count)))
            contact = " | ".join(filter(None, [r[2] or "", r[3] or ""]))
            self.payable_table.setItem(i, 3, QTableWidgetItem(contact))
            pay_btn = QPushButton("付款")
            pay_btn.setObjectName("tableActionOrange")
            pay_btn.clicked.connect(lambda checked, sid=r[0]: self._on_finance_pay(sid))
            self.payable_table.setCellWidget(i, 4, pay_btn)
        self.payable_table.resizeColumnsToContents()

        conn.close()

        # 同时刷新流水记录和筛选对象列表
        self._load_payment_object_filter()
        self.refresh_payment_flow()

    def _load_payment_object_filter(self):
        """加载筛选对象下拉框（客户+上游）"""
        self.payment_object_filter.clear()
        self.payment_object_filter.addItem("全部对象", None)
        customers = get_all_customers()
        for c in customers:
            self.payment_object_filter.addItem(f"客户: {c['name']}", ("customer", c["id"]))
        suppliers = get_all_suppliers()
        for s in suppliers:
            self.payment_object_filter.addItem(f"上游: {s['name']}", ("supplier", s["id"]))

    def refresh_payment_flow(self):
        """刷新收付款流水表格"""
        # 获取筛选条件
        pay_type = self.payment_type_filter.currentText()
        if pay_type == "收款":
            pay_type = "receivable"
        elif pay_type == "付款":
            pay_type = "payable"
        else:
            pay_type = None

        object_data = self.payment_object_filter.currentData()
        customer_id = None
        supplier_id = None
        if object_data:
            if object_data[0] == "customer":
                customer_id = object_data[1]
            elif object_data[0] == "supplier":
                supplier_id = object_data[1]

        date_from = self.payment_date_from.date().toString("yyyy-MM-dd")
        date_to = self.payment_date_to.date().toString("yyyy-MM-dd")

        # 获取数据
        payments = get_all_payments_with_details(pay_type, customer_id, supplier_id, date_from, date_to)

        self.payment_flow_table.setRowCount(len(payments))
        total_receive = 0
        total_pay = 0

        for i, p in enumerate(payments):
            self.payment_flow_table.setItem(i, 0, QTableWidgetItem(str(p["id"])))
            self.payment_flow_table.setItem(i, 1, QTableWidgetItem(p.get("pay_date", "")))

            type_text = "收款" if p.get("type") == "receivable" else "付款"
            type_item = QTableWidgetItem(type_text)
            type_item.setForeground(QColor("#388E3C") if p.get("type") == "receivable" else QColor("#F57C00"))
            self.payment_flow_table.setItem(i, 2, type_item)

            amount_item = QTableWidgetItem(f"¥{p.get('amount', 0):.0f}")
            amount_item.setForeground(QColor("#388E3C") if p.get("type") == "receivable" else QColor("#F57C00"))
            self.payment_flow_table.setItem(i, 3, amount_item)

            self.payment_flow_table.setItem(i, 4, QTableWidgetItem(p.get("method", "")))

            # 关联对象
            obj_name = p.get("customer_name", "") or p.get("supplier_name", "")
            if p.get("type") == "receivable" and obj_name:
                obj_name = f"客户: {obj_name}"
            elif p.get("type") == "payable" and obj_name:
                obj_name = f"上游: {obj_name}"
            self.payment_flow_table.setItem(i, 5, QTableWidgetItem(obj_name))

            self.payment_flow_table.setItem(i, 6, QTableWidgetItem(p.get("remark", "") or ""))

            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)

            edit_btn = QPushButton("修改")
            edit_btn.setObjectName("tableActionPrimary")
            edit_btn.clicked.connect(lambda checked, pid=p["id"]: self._on_edit_payment(pid))
            btn_layout.addWidget(edit_btn)

            del_btn = QPushButton("删除")
            del_btn.setObjectName("tableActionDanger")
            del_btn.clicked.connect(lambda checked, pid=p["id"]: self._on_delete_payment(pid))
            btn_layout.addWidget(del_btn)

            self.payment_flow_table.setCellWidget(i, 7, btn_widget)

            if p.get("type") == "receivable":
                total_receive += p.get("amount", 0)
            else:
                total_pay += p.get("amount", 0)

        self.payment_flow_table.resizeColumnsToContents()
        self.payment_flow_stats.setText(
            f"共 {len(payments)} 条记录 | 收款合计: ¥{total_receive:.0f} | 付款合计: ¥{total_pay:.0f}"
        )

    def _on_edit_payment(self, payment_id):
        """修改收付款记录（两次确认）"""
        payment = get_payment_by_id(payment_id)
        if not payment:
            QMessageBox.warning(self, "错误", "无法获取记录信息")
            return

        # 第一次确认：显示原记录信息
        type_text = "收款" if payment["type"] == "receivable" else "付款"
        obj_name = payment.get("customer_name", "") or payment.get("supplier_name", "")
        info = f"原记录信息:\n\n类型: {type_text}\n金额: ¥{payment['amount']:.0f}\n日期: {payment['pay_date']}\n方式: {payment['method']}\n关联对象: {obj_name}\n备注: {payment['remark'] or '无'}\n\n是否要修改此记录？"
        reply = QMessageBox.question(self, "确认修改 - 第一步", info, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return

        # 弹出编辑对话框
        dlg = PaymentEditDialog(self, payment)
        if dlg.exec():
            new_data = dlg.get_data()
            if new_data["amount"] <= 0:
                QMessageBox.warning(self, "提示", "金额必须大于 0")
                return

            # 第二次确认：显示修改内容对比
            changes = []
            if payment["amount"] != new_data["amount"]:
                changes.append(f"金额: ¥{payment['amount']:.0f} → ¥{new_data['amount']:.0f}")
            if payment["pay_date"] != new_data["pay_date"]:
                changes.append(f"日期: {payment['pay_date']} → {new_data['pay_date']}")
            if payment["method"] != new_data["method"]:
                changes.append(f"方式: {payment['method']} → {new_data['method']}")
            if (payment["remark"] or "") != new_data["remark"]:
                changes.append(f"备注: {payment['remark'] or '无'} → {new_data['remark'] or '无'}")

            if not changes:
                QMessageBox.information(self, "提示", "没有任何修改")
                return

            confirm_text = f"确认以下修改:\n\n" + "\n".join(changes) + "\n\n修改后将自动更新关联数据（报价已收金额/供应商欠款）。\n是否确认修改？"
            reply2 = QMessageBox.question(self, "确认修改 - 第二步", confirm_text, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply2 != QMessageBox.StandardButton.Yes:
                return

            # 执行修改
            success, msg = update_payment(payment_id, new_data["amount"], new_data["pay_date"], new_data["method"], new_data["remark"])
            if success:
                QMessageBox.information(self, "成功", "修改成功！关联数据已同步更新。")
                self.refresh_payment_flow()
                self.refresh_finance()
                self.refresh_records()
            else:
                QMessageBox.warning(self, "失败", msg)

    def _on_delete_payment(self, payment_id):
        """删除收付款记录（两次确认 + 回滚）"""
        payment = get_payment_by_id(payment_id)
        if not payment:
            QMessageBox.warning(self, "错误", "无法获取记录信息")
            return

        # 第一次确认：显示记录信息
        type_text = "收款" if payment["type"] == "receivable" else "付款"
        obj_name = payment.get("customer_name", "") or payment.get("supplier_name", "")
        info = f"即将删除的记录:\n\n类型: {type_text}\n金额: ¥{payment['amount']:.0f}\n日期: {payment['pay_date']}\n方式: {payment['method']}\n关联对象: {obj_name}\n备注: {payment['remark'] or '无'}\n\n删除后将自动回滚关联数据。\n是否继续？"
        reply = QMessageBox.question(self, "确认删除 - 第一步", info, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return

        # 第二次确认：显示影响范围
        if payment["type"] == "receivable" and payment.get("quote_id"):
            impact = f"影响范围:\n• 报价记录 #{payment['quote_id']} 的已收金额将减少 ¥{payment['amount']:.0f}\n• 若已收金额不足，订单状态可能从「已收款」回退为「已出库」"
        elif payment["type"] == "payable" and payment.get("supplier_id"):
            impact = f"影响范围:\n• 上游供应商「{obj_name}」的欠款将增加 ¥{payment['amount']:.0f}"
        else:
            impact = "此记录无直接关联数据，删除后仅影响流水统计。"

        confirm_text = f"⚠️ 最终确认\n\n{impact}\n\n此操作不可撤销！\n是否确认删除？"
        reply2 = QMessageBox.question(self, "确认删除 - 第二步", confirm_text, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply2 != QMessageBox.StandardButton.Yes:
            return

        # 执行删除
        success, msg, affected = delete_payment(payment_id)
        if success:
            QMessageBox.information(self, "成功", "删除成功！关联数据已回滚。")
            self.refresh_payment_flow()
            self.refresh_finance()
            self.refresh_records()
        else:
            QMessageBox.warning(self, "失败", msg)

    def _on_finance_receive(self, customer_id):
        """收款操作（带预览确认）"""
        from src.models.database import get_connection
        conn = get_connection()
        unpaid = conn.execute("""
            SELECT id, quote_price, quote_quantity, received_amount,
                   quote_price * quote_quantity - received_amount as remaining
            FROM quotes
            WHERE customer_id=? AND status IN ('已报价','已出库','待确认')
            AND quote_price * quote_quantity > received_amount
            ORDER BY quote_date ASC
        """, (customer_id,)).fetchall()
        total_pending = sum(r[4] for r in unpaid)
        conn.close()

        if not unpaid:
            QMessageBox.information(self, "提示", "该客户没有待收款项")
            return

        # 弹出收款对话框（带预览）
        dlg = PaymentDialog(self, title="收款", pay_type="receivable", preview_pending=total_pending)
        if dlg.exec():
            data = dlg.get_data()
            if data["amount"] <= 0:
                QMessageBox.warning(self, "提示", "请输入收款金额")
                return

            # 预览确认：显示收款后的剩余待收
            remaining_after = total_pending - data["amount"]
            preview_text = f"收款确认:\n\n本次收款: ¥{data['amount']:.0f}\n当前待收: ¥{total_pending:.0f}\n收款后剩余待收: ¥{remaining_after:.0f}\n\n收款方式: {data['method']}\n收款日期: {data['pay_date']}"
            if remaining_after < 0:
                preview_text += "\n\n⚠️ 注意: 收款金额超过待收金额，将自动结清所有欠款。"

            reply = QMessageBox.question(self, "确认收款", preview_text, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return

            # 批量收款在同一个事务中执行，保证原子性
            conn = get_connection()
            try:
                remaining = data["amount"]
                for q in unpaid:
                    if remaining <= 0:
                        break
                    qid, _, _, _, pending = q
                    apply_amount = min(remaining, pending)
                    _add_payment_raw(
                        conn,
                        quote_id=qid,
                        customer_id=customer_id,
                        pay_type="receivable",
                        amount=apply_amount,
                        pay_date=data["pay_date"],
                        method=data["method"],
                        remark=data["remark"],
                    )
                    remaining -= apply_amount
                conn.commit()
                QMessageBox.information(self, "成功", f"收款 ¥{data['amount']:.2f} 已记录！")
            except Exception as e:
                conn.rollback()
                QMessageBox.critical(self, "收款失败", f"收款操作失败: {str(e)}")
                return
            finally:
                conn.close()
            self.refresh_finance()
            self.refresh_records()

    def _on_finance_pay(self, supplier_id):
        """付款操作"""
        dlg = PaymentDialog(self, title="付款", pay_type="payable")
        if dlg.exec():
            data = dlg.get_data()
            if data["amount"] <= 0:
                QMessageBox.warning(self, "提示", "请输入付款金额")
                return
            add_payment(
                supplier_id=supplier_id,
                pay_type="payable",
                amount=data["amount"],
                pay_date=data["pay_date"],
                method=data["method"],
                remark=data["remark"],
            )
            QMessageBox.information(self, "成功", f"付款 ¥{data['amount']:.2f} 已记录！")
            self.refresh_finance()

    # -------------------------------------------------------
    # Tab 构建
    # -------------------------------------------------------
    def _build_customer_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("+ 新增客户")
        add_btn.setObjectName("primaryBtn")
        add_btn.clicked.connect(self.on_add_customer)
        del_btn = QPushButton("删除客户")
        del_btn.setObjectName("dangerBtn")
        del_btn.clicked.connect(self.on_delete_customer)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(del_btn)
        btn_row.addStretch()

        self.customer_search = QLineEdit()
        self.customer_search.setObjectName("globalSearch")
        self.customer_search.setPlaceholderText("搜索客户...")
        self.customer_search.textChanged.connect(self.refresh_customer_list)
        btn_row.addWidget(self.customer_search)
        layout.addLayout(btn_row)

        self.customer_table = QTableWidget()
        self.customer_table.setAlternatingRowColors(True)
        self.customer_table.setColumnCount(6)
        self.customer_table.setHorizontalHeaderLabels(["ID", "名称", "微信", "QQ", "电话", "备注"])
        self.customer_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.customer_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.customer_table.setColumnHidden(0, True)
        self.customer_table.horizontalHeader().setStretchLastSection(True)
        self.customer_table.doubleClicked.connect(self.on_edit_customer_from_table)
        self.customer_table.cellClicked.connect(self.on_customer_cell_clicked)
        layout.addWidget(self.customer_table)

        history_group = QGroupBox("购买历史")
        history_layout = QVBoxLayout(history_group)

        self.customer_stats_label = QLabel("请选择客户查看购买历史")
        self.customer_stats_label.setObjectName("summaryLabel")
        history_layout.addWidget(self.customer_stats_label)

        self.customer_history_table = QTableWidget()
        self.customer_history_table.setAlternatingRowColors(True)
        self.customer_history_table.setColumnCount(8)
        self.customer_history_table.setHorizontalHeaderLabels(["日期", "机型", "CPU", "数量", "购入价", "报价", "毛利", "备注"])
        self.customer_history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.customer_history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.customer_history_table.horizontalHeader().setStretchLastSection(True)
        self.customer_history_table.setColumnWidth(7, 120)
        history_layout.addWidget(self.customer_history_table)

        layout.addWidget(history_group)

        return tab

    def _build_records_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        btn_row = QHBoxLayout()
        self.confirm_record_btn = QPushButton("确认报价")
        self.confirm_record_btn.setObjectName("successBtn")
        self.confirm_record_btn.clicked.connect(self.on_confirm_quote)
        self.ship_record_btn = QPushButton("出库")
        self.ship_record_btn.setObjectName("warningBtn")
        self.ship_record_btn.clicked.connect(self.on_ship_quote)
        self.receive_btn = QPushButton("收款")
        self.receive_btn.setObjectName("primaryBtn")
        self.receive_btn.clicked.connect(self.on_receive_payment)
        self.cancel_record_btn = QPushButton("取消订单")
        self.cancel_record_btn.setObjectName("ghostBtn")
        self.cancel_record_btn.clicked.connect(self.on_cancel_quote)
        self.edit_record_btn = QPushButton("编辑")
        self.edit_record_btn.setObjectName("ghostBtn")
        self.edit_record_btn.clicked.connect(self.on_edit_quote)
        self.del_record_btn = QPushButton("删除")
        self.del_record_btn.setObjectName("dangerBtn")
        self.del_record_btn.clicked.connect(self.on_delete_quote)
        self.refresh_records_btn = QPushButton("刷新")
        self.refresh_records_btn.setObjectName("ghostBtn")
        self.refresh_records_btn.clicked.connect(self.refresh_records)
        btn_row.addWidget(self.confirm_record_btn)
        btn_row.addWidget(self.ship_record_btn)
        btn_row.addWidget(self.receive_btn)
        btn_row.addWidget(self.cancel_record_btn)
        btn_row.addWidget(self.edit_record_btn)
        btn_row.addWidget(self.del_record_btn)
        btn_row.addWidget(self.refresh_records_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.record_search = QLineEdit()
        self.record_search.setPlaceholderText("搜索机型/序列...")
        self.record_search.textChanged.connect(self.refresh_records)

        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.dateChanged.connect(self.refresh_records)

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.dateChanged.connect(self.refresh_records)

        self.status_filter = QComboBox()
        self.status_filter.addItems(["全部状态", "待确认", "已报价", "已出库", "已收款", "已取消"])
        self.status_filter.currentTextChanged.connect(self.refresh_records)

        self.export_records_btn = QPushButton("导出 Excel")
        self.export_records_btn.setObjectName("ghostBtn")
        self.export_records_btn.clicked.connect(self.on_export_records_excel)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("开始:"))
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(QLabel("结束:"))
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(QLabel("状态:"))
        filter_layout.addWidget(self.status_filter)
        filter_layout.addWidget(QLabel("搜索:"))
        filter_layout.addWidget(self.record_search)
        filter_layout.addStretch()
        filter_layout.addWidget(self.export_records_btn)

        filter_card = QFrame()
        filter_card.setObjectName("filterCard")
        filter_card_layout = QVBoxLayout(filter_card)
        filter_card_layout.setContentsMargins(12, 8, 12, 8)
        filter_card_layout.addLayout(filter_layout)
        layout.addWidget(filter_card)

        self.record_table = QTableWidget()
        self.record_table.setAlternatingRowColors(True)
        self.record_table.setColumnCount(17)
        self.record_table.setHorizontalHeaderLabels(
            ["ID", "日期", "客户", "机型", "CPU", "内存", "硬盘", "显卡", "上游",
             "购入价", "数量", "对外报价", "状态", "已收款", "SN", "备注", "打款"]
        )
        self.record_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.record_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.record_table.horizontalHeader().setStretchLastSection(True)
        self.record_table.setColumnHidden(0, True)
        layout.addWidget(self.record_table)

        self.stats_label = QLabel()
        self.stats_label.setObjectName("summaryLabel")
        layout.addWidget(self.stats_label)

        return tab

    # -------------------------------------------------------
    # 机型管理
    # -------------------------------------------------------
    def refresh_product_list(self, keyword=None):
        if keyword:
            products = search_products(keyword)
        else:
            products = get_all_products()

        self.product_table.setRowCount(len(products))
        for i, p in enumerate(products):
            self.product_table.setItem(i, 0, QTableWidgetItem(str(p["id"])))
            self.product_table.setItem(i, 1, QTableWidgetItem(p.get("series", "")))
            self.product_table.setItem(i, 2, QTableWidgetItem(p.get("cpu", "")))
            self.product_table.setItem(i, 3, QTableWidgetItem(p.get("ram", "")))
            self.product_table.setItem(i, 4, QTableWidgetItem(p.get("storage", "")))
            self.product_table.setItem(i, 5, QTableWidgetItem(p.get("gpu", "")))
            self.product_table.setItem(i, 6, QTableWidgetItem(p.get("screen", "")))
            self.product_table.setItem(i, 7, QTableWidgetItem(p.get("note", "")))
            remaining = get_total_remaining(p["id"])
            self.product_table.setItem(i, 8, QTableWidgetItem(str(remaining)))

        self.product_table.resizeColumnsToContents()
        self.product_table.setColumnWidth(1, 180)
        self.status_label.setText(f"共 {len(products)} 条机型")

    def on_search(self, text):
        self.refresh_product_list(text.strip())

    def on_product_selected(self):
        row = self.product_table.currentRow()
        if row < 0:
            return
        pid = int(self.product_table.item(row, 0).text())
        series = self.product_table.item(row, 1).text()
        cpu = self.product_table.item(row, 2).text()
        ram = self.product_table.item(row, 3).text()
        storage = self.product_table.item(row, 4).text()
        gpu = self.product_table.item(row, 5).text()
        screen = self.product_table.item(row, 6).text()
        note_item = self.product_table.item(row, 7)
        note = note_item.text() if note_item else ""
        self.quote_panel.load_product(pid, series, cpu, ram, storage, gpu, screen, note)

    def on_add_product(self):
        dlg = ProductEditDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            if not data["series"]:
                QMessageBox.warning(self, "提示", "系列名称不能为空")
                return
            add_product(**data)
            add_operation_log("新增机型", "products", 0, f"系列={data['series']}")
            self.refresh_product_list(self.search_edit.text().strip())

    def on_edit_product(self):
        row = self.product_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个机型")
            return
        pid = int(self.product_table.item(row, 0).text())
        note_item = self.product_table.item(row, 7)
        current = {
            "series": self.product_table.item(row, 1).text(),
            "cpu": self.product_table.item(row, 2).text(),
            "ram": self.product_table.item(row, 3).text(),
            "storage": self.product_table.item(row, 4).text(),
            "gpu": self.product_table.item(row, 5).text(),
            "screen": self.product_table.item(row, 6).text(),
            "note": note_item.text() if note_item else "",
        }
        dlg = ProductEditDialog(self, current)
        if dlg.exec():
            data = dlg.get_data()
            if not data["series"]:
                QMessageBox.warning(self, "提示", "系列名称不能为空")
                return
            update_product(pid, **data)
            self.refresh_product_list(self.search_edit.text().strip())

    def on_delete_product(self):
        row = self.product_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个机型")
            return
        pid = int(self.product_table.item(row, 0).text())
        series = self.product_table.item(row, 1).text()
        reply = QMessageBox.question(
            self, "确认删除", f"确定删除机型「{series}」及其所有库存记录？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            delete_product(pid)
            add_operation_log("删除机型", "products", pid, f"系列={series}")
            self.refresh_product_list(self.search_edit.text().strip())

    # -------------------------------------------------------
    # 客户管理
    # -------------------------------------------------------
    def refresh_customer_list(self):
        keyword = self.customer_search.text().strip()
        customers = search_customers(keyword) if keyword else get_all_customers()
        self.customer_table.setRowCount(len(customers))
        for i, c in enumerate(customers):
            self.customer_table.setItem(i, 0, QTableWidgetItem(str(c["id"])))
            self.customer_table.setItem(i, 1, QTableWidgetItem(c["name"]))
            self.customer_table.setItem(i, 2, QTableWidgetItem(c.get("wechat", "")))
            self.customer_table.setItem(i, 3, QTableWidgetItem(c.get("qq", "")))
            self.customer_table.setItem(i, 4, QTableWidgetItem(c.get("phone", "")))
            self.customer_table.setItem(i, 5, QTableWidgetItem(c.get("note", "")))
        self.customer_table.resizeColumnsToContents()

    def on_add_customer(self):
        dlg = CustomerDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                QMessageBox.warning(self, "提示", "客户名称不能为空")
                return
            add_customer(**data)
            add_operation_log("新增客户", "customers", 0, f"名称={data['name']}")
            self.refresh_customer_list()

    def on_customer_cell_clicked(self, row, col):
        if row < 0:
            self.customer_stats_label.setText("请选择客户查看购买历史")
            self.customer_history_table.setRowCount(0)
            return
        
        cid = int(self.customer_table.item(row, 0).text())
        customer_name = self.customer_table.item(row, 1).text()
        
        from src.models.database import get_customer_quotes, get_customer_stats
        quotes = get_customer_quotes(cid)
        stats = get_customer_stats(cid)
        
        self.customer_stats_label.setText(
            f"客户: {customer_name} | 总成交: {stats['total_quotes']}单 | 总金额: ¥{stats['total_amount']:.0f} | 总毛利: ¥{stats['total_profit']:.0f}"
        )
        
        self.customer_history_table.setRowCount(len(quotes))
        for i, q in enumerate(quotes):
            purchase_price = q.get("purchase_price", 0) or 0
            quote_price = q.get("quote_price", 0) or 0
            quantity = q.get("quote_quantity", 1) or 1
            profit = (quote_price - purchase_price) * quantity
            
            self.customer_history_table.setItem(i, 0, QTableWidgetItem(q.get("quote_date", "")))
            self.customer_history_table.setItem(i, 1, QTableWidgetItem(q.get("series", "")))
            self.customer_history_table.setItem(i, 2, QTableWidgetItem(q.get("cpu", "")))
            self.customer_history_table.setItem(i, 3, QTableWidgetItem(str(quantity)))
            self.customer_history_table.setItem(i, 4, QTableWidgetItem(f"¥{purchase_price:.0f}"))
            self.customer_history_table.setItem(i, 5, QTableWidgetItem(f"¥{quote_price:.0f}"))
            self.customer_history_table.setItem(i, 6, QTableWidgetItem(f"¥{profit:.0f}"))
            self.customer_history_table.setItem(i, 7, QTableWidgetItem(q.get("remark", "")))
        
        self.customer_history_table.resizeColumnsToContents()

    def on_edit_customer_from_table(self):
        row = self.customer_table.currentRow()
        if row < 0:
            return
        cid = int(self.customer_table.item(row, 0).text())
        current = {
            "name": self.customer_table.item(row, 1).text(),
            "wechat": self.customer_table.item(row, 2).text() if self.customer_table.item(row, 2) else "",
            "qq": self.customer_table.item(row, 3).text() if self.customer_table.item(row, 3) else "",
            "phone": self.customer_table.item(row, 4).text() if self.customer_table.item(row, 4) else "",
            "note": "",
        }
        dlg = CustomerDialog(self, current)
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                return
            from src.models.database import update_product  # reuse; add customer update
            # 简单实现：删除重建
            from src.models.database import get_connection
            conn = get_connection()
            conn.execute(
                "UPDATE customers SET name=?, wechat=?, qq=?, phone=?, note=? WHERE id=?",
                (data["name"], data["wechat"], data["qq"], data["phone"], data["note"], cid),
            )
            conn.commit()
            conn.close()
            self.refresh_customer_list()

    def on_delete_customer(self):
        row = self.customer_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个客户")
            return
        cid = int(self.customer_table.item(row, 0).text())
        name = self.customer_table.item(row, 1).text()
        from src.models.database import get_connection
        conn = get_connection()
        quote_count = conn.execute("SELECT COUNT(*) FROM quotes WHERE customer_id=?", (cid,)).fetchone()[0]
        conn.close()
        if quote_count > 0:
            reply = QMessageBox.question(
                self, "确认删除",
                f"客户「{name}」有 {quote_count} 条关联报价记录。\n\n确定删除该客户及其所有关联数据？\n此操作不可恢复。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
        else:
            reply = QMessageBox.question(
                self, "确认删除", f"确定删除客户「{name}」？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
        if reply == QMessageBox.StandardButton.Yes:
            delete_customer_cascade(cid)
            add_operation_log("删除客户", "customers", cid, f"名称={name}")
            self.refresh_customer_list()
            self.refresh_records()

    # -------------------------------------------------------
    # 上游管理
    # -------------------------------------------------------
    def _build_supplier_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("+ 新增上游")
        add_btn.setObjectName("primaryBtn")
        add_btn.clicked.connect(self.on_add_supplier)
        del_btn = QPushButton("删除上游")
        del_btn.setObjectName("dangerBtn")
        del_btn.clicked.connect(self.on_delete_supplier)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(del_btn)
        btn_row.addStretch()

        self.supplier_search = QLineEdit()
        self.supplier_search.setObjectName("globalSearch")
        self.supplier_search.setPlaceholderText("搜索上游...")
        self.supplier_search.textChanged.connect(self.refresh_supplier_list)
        btn_row.addWidget(self.supplier_search)
        layout.addLayout(btn_row)

        self.supplier_table = QTableWidget()
        self.supplier_table.setAlternatingRowColors(True)
        self.supplier_table.setColumnCount(6)
        self.supplier_table.setHorizontalHeaderLabels(["ID", "名称", "微信", "QQ", "电话", "备注"])
        self.supplier_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.supplier_table.setColumnHidden(0, True)
        self.supplier_table.horizontalHeader().setStretchLastSection(True)
        self.supplier_table.cellClicked.connect(self.on_supplier_cell_clicked)
        self.supplier_table.doubleClicked.connect(self.on_edit_supplier_from_table)
        layout.addWidget(self.supplier_table)

        history_group = QGroupBox("采购历史")
        history_layout = QVBoxLayout(history_group)

        self.supplier_stats_label = QLabel("请选择上游查看采购历史")
        self.supplier_stats_label.setObjectName("summaryLabel")
        history_layout.addWidget(self.supplier_stats_label)

        self.supplier_history_table = QTableWidget()
        self.supplier_history_table.setAlternatingRowColors(True)
        self.supplier_history_table.setColumnCount(7)
        self.supplier_history_table.setHorizontalHeaderLabels(["入库日期", "机型", "CPU", "数量", "购入价", "总金额", "备注"])
        self.supplier_history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.supplier_history_table.horizontalHeader().setStretchLastSection(True)
        history_layout.addWidget(self.supplier_history_table)

        layout.addWidget(history_group)

        return tab

    def on_supplier_cell_clicked(self, row, col):
        if row < 0:
            self.supplier_stats_label.setText("请选择上游查看采购历史")
            self.supplier_history_table.setRowCount(0)
            return

        sid = int(self.supplier_table.item(row, 0).text())
        supplier_name = self.supplier_table.item(row, 1).text()

        from src.models.database import get_connection
        conn = get_connection()
        rows = conn.execute("""
            SELECT b.date, p.series, p.cpu, b.quantity, b.purchase_price, b.remark
            FROM batches b
            JOIN products p ON b.product_id = p.id
            WHERE b.supplier_id = ?
            ORDER BY b.date DESC, b.id DESC
        """, (sid,)).fetchall()
        conn.close()

        total_amount = 0
        self.supplier_history_table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            total_amount += (r[4] or 0) * (r[3] or 0)
            self.supplier_history_table.setItem(i, 0, QTableWidgetItem(r[0] or ""))
            self.supplier_history_table.setItem(i, 1, QTableWidgetItem(r[1] or ""))
            self.supplier_history_table.setItem(i, 2, QTableWidgetItem(r[2] or ""))
            self.supplier_history_table.setItem(i, 3, QTableWidgetItem(str(r[3] or 0)))
            self.supplier_history_table.setItem(i, 4, QTableWidgetItem(f"¥{(r[4] or 0):.0f}"))
            self.supplier_history_table.setItem(i, 5, QTableWidgetItem(f"¥{(r[4] or 0) * (r[3] or 0):.0f}"))
            self.supplier_history_table.setItem(i, 6, QTableWidgetItem(r[5] or ""))
        self.supplier_history_table.resizeColumnsToContents()

        self.supplier_stats_label.setText(
            f"上游: {supplier_name} | 总批次数: {len(rows)} | 总金额: ¥{total_amount:.0f}"
        )

    def refresh_supplier_list(self):
        keyword = self.supplier_search.text().strip()
        suppliers = search_suppliers(keyword) if keyword else get_all_suppliers()
        self.supplier_table.setRowCount(len(suppliers))
        for i, s in enumerate(suppliers):
            self.supplier_table.setItem(i, 0, QTableWidgetItem(str(s["id"])))
            self.supplier_table.setItem(i, 1, QTableWidgetItem(s["name"]))
            self.supplier_table.setItem(i, 2, QTableWidgetItem(s.get("wechat", "")))
            self.supplier_table.setItem(i, 3, QTableWidgetItem(s.get("qq", "")))
            self.supplier_table.setItem(i, 4, QTableWidgetItem(s.get("phone", "")))
            self.supplier_table.setItem(i, 5, QTableWidgetItem(s.get("note", "")))
        self.supplier_table.resizeColumnsToContents()

    def on_add_supplier(self):
        dlg = CustomerDialog(self)
        dlg.setWindowTitle("新增上游")
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                QMessageBox.warning(self, "提示", "上游名称不能为空")
                return
            add_supplier(**data)
            add_operation_log("新增供应商", "suppliers", 0, f"名称={data['name']}")
            self.refresh_supplier_list()

    def on_edit_supplier_from_table(self):
        row = self.supplier_table.currentRow()
        if row < 0:
            return
        sid = int(self.supplier_table.item(row, 0).text())
        current = {
            "name": self.supplier_table.item(row, 1).text(),
            "wechat": self.supplier_table.item(row, 2).text() if self.supplier_table.item(row, 2) else "",
            "qq": self.supplier_table.item(row, 3).text() if self.supplier_table.item(row, 3) else "",
            "phone": self.supplier_table.item(row, 4).text() if self.supplier_table.item(row, 4) else "",
            "note": "",
        }
        dlg = CustomerDialog(self, current)
        dlg.setWindowTitle("编辑上游")
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                return
            from src.models.database import get_connection
            conn = get_connection()
            conn.execute(
                "UPDATE suppliers SET name=?, wechat=?, qq=?, phone=?, note=? WHERE id=?",
                (data["name"], data["wechat"], data["qq"], data["phone"], data["note"], sid),
            )
            conn.commit()
            conn.close()
            self.refresh_supplier_list()

    def on_delete_supplier(self):
        row = self.supplier_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个上游")
            return
        sid = int(self.supplier_table.item(row, 0).text())
        name = self.supplier_table.item(row, 1).text()
        from src.models.database import get_connection
        conn = get_connection()
        batch_count = conn.execute("SELECT COUNT(*) FROM batches WHERE supplier_id=?", (sid,)).fetchone()[0]
        payment_count = conn.execute("SELECT COUNT(*) FROM payments WHERE supplier_id=?", (sid,)).fetchone()[0]
        conn.close()
        if batch_count > 0 or payment_count > 0:
            reply = QMessageBox.question(
                self, "确认删除",
                f"上游「{name}」有 {batch_count} 条关联批次和 {payment_count} 条付款记录。\n\n"
                f"删除后关联批次的上游将设为空，付款记录将被删除。\n此操作不可恢复。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
        else:
            reply = QMessageBox.question(
                self, "确认删除", f"确定删除上游「{name}」？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
        if reply == QMessageBox.StandardButton.Yes:
            delete_supplier_cascade(sid)
            add_operation_log("删除供应商", "suppliers", sid, f"名称={name}")
            self.refresh_supplier_list()
            self.refresh_records()

    # -------------------------------------------------------
    # 报价记录
    # -------------------------------------------------------
    def refresh_records(self):
        keyword = self.record_search.text().strip()
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        quotes = search_quotes(keyword, date_from, date_to)

        status_filter = self.status_filter.currentText()
        if status_filter != "全部状态":
            quotes = [q for q in quotes if q.get("status", "待确认") == status_filter]

        STATUS_COLORS = {
            "待确认": "#9E9E9E",
            "已报价": "#1976D2",
            "已出库": "#F57C00",
            "已收款": "#388E3C",
            "已取消": "#BDBDBD",
        }

        self.record_table.setSortingEnabled(False)
        self.record_table.setRowCount(len(quotes))
        total_cost = 0
        total_sale = 0
        for i, q in enumerate(quotes):
            status = q.get("status", "待确认")
            received = q.get("received_amount", 0) or 0
            sn_list = q.get("sn_list", "") or q.get("batch_sn_list", "") or ""
            quote_price = q.get("quote_price", 0) or 0
            quote_quantity = q.get("quote_quantity", 1) or 1
            total_amount = quote_price * quote_quantity

            self.record_table.setItem(i, 0, QTableWidgetItem(str(q.get("id", ""))))
            self.record_table.setItem(i, 1, QTableWidgetItem(q.get("quote_date", "")))
            self.record_table.setItem(i, 2, QTableWidgetItem(q.get("customer_name", "")))
            self.record_table.setItem(i, 3, QTableWidgetItem(q.get("series", "")))
            self.record_table.setItem(i, 4, QTableWidgetItem(q.get("cpu", "")))
            self.record_table.setItem(i, 5, QTableWidgetItem(q.get("ram", "")))
            self.record_table.setItem(i, 6, QTableWidgetItem(q.get("storage", "")))
            self.record_table.setItem(i, 7, QTableWidgetItem(q.get("gpu", "")))
            self.record_table.setItem(i, 8, QTableWidgetItem(q.get("supplier_name", "") or ""))
            self.record_table.setItem(i, 9, QTableWidgetItem(f"¥{q.get('purchase_price', 0):.0f}" if q.get('purchase_price') else ""))
            self.record_table.setItem(i, 10, QTableWidgetItem(str(quote_quantity)))
            self.record_table.setItem(i, 11, QTableWidgetItem(f"¥{quote_price:.0f}" if quote_price else ""))

            status_item = QTableWidgetItem(status)
            color = STATUS_COLORS.get(status, "#333")
            status_item.setForeground(Qt.GlobalColor.white)
            status_item.setBackground(QColor(color))
            self.record_table.setItem(i, 12, status_item)

            received_text = f"¥{received:.0f}" if received > 0 else "¥0"
            received_item = QTableWidgetItem(received_text)
            if received >= total_amount and total_amount > 0:
                received_item.setForeground(QColor("#388E3C"))
            self.record_table.setItem(i, 13, received_item)

            self.record_table.setItem(i, 14, QTableWidgetItem(sn_list))

            batch_remark = q.get("batch_remark", "") or ""
            quote_remark = q.get("remark", "") or ""
            merged_remark = " | ".join(filter(None, [batch_remark, quote_remark]))
            self.record_table.setItem(i, 15, QTableWidgetItem(merged_remark))
            self.record_table.setItem(i, 16, QTableWidgetItem(q.get("paid", "否")))

            total_cost += (q.get("purchase_price", 0) or 0) * quote_quantity
            total_sale += quote_price * quote_quantity

            # 收款提醒：已出库超过7天未收满的订单，整行红色高亮
            if status == "已出库" and received < total_amount and total_amount > 0:
                from datetime import date, datetime
                quote_date = q.get("quote_date", "")
                if quote_date:
                    try:
                        quote_dt = datetime.strptime(quote_date, "%Y-%m-%d").date()
                        if (date.today() - quote_dt).days > 7:
                            for col in range(self.record_table.columnCount()):
                                item = self.record_table.item(i, col)
                                if item:
                                    item.setForeground(QColor("#D32F2F"))
                    except ValueError:
                        pass

        self.record_table.setSortingEnabled(True)
        self.record_table.resizeColumnsToContents()
        self.record_table.setColumnWidth(12, 70)
        self.record_table.setColumnWidth(13, 80)
        self.record_table.setColumnWidth(14, 120)
        profit = total_sale - total_cost
        self.stats_label.setText(
            f"共 {len(quotes)} 条记录  |  总购入: ¥{total_cost:.0f}  |  总报价: ¥{total_sale:.0f}  |  毛利: ¥{profit:.0f}"
        )

    def on_edit_quote(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        quote = get_quote_by_id(quote_id)
        if not quote:
            QMessageBox.warning(self, "错误", "无法获取报价记录信息")
            return
        dlg = QuoteEditDialog(self, quote)
        if dlg.exec():
            data = dlg.get_data()
            if data["batch_id"]:
                update_quote(
                    quote_id,
                    data["batch_id"],
                    data["customer_id"],
                    data["quote_price"],
                    data["quote_quantity"],
                    data["quote_date"],
                    data["remark"],
                    data["paid"],
                    data.get("sn_list", ""),
                )
                self.refresh_records()

    def on_delete_quote(self):
        row = self.record_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一条报价记录")
            return
        quote_id = int(self.record_table.item(row, 0).text())
        series = self.record_table.item(row, 3).text() if self.record_table.item(row, 3) else ""
        customer = self.record_table.item(row, 2).text() if self.record_table.item(row, 2) else ""
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定删除报价记录「{customer} - {series}」？\n此操作不可恢复。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            delete_quote(quote_id)
            add_operation_log("删除报价", "quotes", quote_id, f"客户={customer}, 机型={series}")
            self.refresh_records()

    # -------------------------------------------------------
    # 导入 Word
    # -------------------------------------------------------
    def on_import_word(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "选择 Word 价格表", "", "Word 文档 (*.docx *.doc)"
        )
        if not filepath:
            return

        try:
            products = parse_word_pricelist(filepath)
        except Exception as e:
            QMessageBox.critical(self, "解析失败", f"无法解析 Word 文件:\n{str(e)}")
            return

        if not products:
            QMessageBox.warning(self, "提示", "未从文档中识别到任何机型")
            return

        # 预览确认
        preview = "\n".join([f"{p.get('series','?')} | {p.get('cpu','')} | {p.get('ram','')} | {p.get('storage','')} | {p.get('gpu','')}" for p in products[:20]])
        if len(products) > 20:
            preview += f"\n... 及其他 {len(products)-20} 条"

        reply = QMessageBox.question(
            self, "确认导入",
            f"识别到 {len(products)} 条机型，预览如下:\n\n{preview}\n\n是否导入？"
            "\n（已存在的机型将被跳过）",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        # 导入
        imported = 0
        for p in products:
            series = p.get("series", "")
            if not series or series == "未识别":
                continue
            # 检查是否已存在（按全部字段去重）
            existing = search_products(series)
            skip = False
            for e in existing:
                if (e["series"] == series and
                    e.get("cpu", "") == p.get("cpu", "") and
                    e.get("ram", "") == p.get("ram", "") and
                    e.get("storage", "") == p.get("storage", "") and
                    e.get("gpu", "") == p.get("gpu", "") and
                    e.get("screen", "") == p.get("screen", "") and
                    e.get("note", "") == p.get("note", "")):
                    skip = True
                    break
            if not skip:
                add_product(
                    series=series,
                    cpu=p.get("cpu", ""),
                    ram=p.get("ram", ""),
                    storage=p.get("storage", ""),
                    gpu=p.get("gpu", ""),
                    screen=p.get("screen", ""),
                    note=p.get("note", ""),
                )
                imported += 1

        self.refresh_product_list(self.search_edit.text().strip())
        QMessageBox.information(self, "导入完成", f"成功导入 {imported} 条新机型\n跳过 {len(products)-imported} 条已存在的记录")

        # ---- Skill 2: 价格异动哨兵 ----
        # 保存本次快照
        try:
            snapshot_id, count = save_snapshot(products)
        except Exception:
            snapshot_id = None

        # 和上一次快照做对比
        if snapshot_id:
            try:
                today = datetime.now().strftime("%Y-%m-%d")
                old_snapshot = get_latest_snapshot(before_date=today)
                # 如果今天之前没有更早的快照，再取最新的（可能是今天刚导入的前一次）
                if old_snapshot is None:
                    old_snapshot = get_latest_snapshot()
                    # 排除当前这次刚保存的
                    if old_snapshot and old_snapshot["snapshot_id"] == snapshot_id:
                        old_snapshot = None

                if old_snapshot:
                    diff = diff_snapshots(old_snapshot, products)
                    if diff["added"] or diff["removed"]:
                        self._show_diff_report(diff)
                else:
                    QMessageBox.information(
                        self, "首次快照",
                        f"已保存价格快照（{count} 条机型）。\n下次导入时将自动对比变动。"
                    )
            except Exception as e:
                pass  # 异动分析失败不影响主流程

    # -------------------------------------------------------
    # Skill 3: 智能跟单提醒
    # -------------------------------------------------------
    def on_follow_up(self):
        stale = get_stale_quotes(stale_days=3)
        text = format_reminder_text(stale)

        dlg = QDialog(self)
        dlg.setWindowTitle("🔔 跟单提醒")
        dlg.setMinimumSize(500, 400)
        layout = QVBoxLayout(dlg)

        label = QLabel(text)
        label.setObjectName("reportText")
        label.setWordWrap(True)
        layout.addWidget(label)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dlg.accept)
        layout.addWidget(buttons)
        dlg.exec()

    # -------------------------------------------------------
    # Skill 4: 月度经营报告
    # -------------------------------------------------------
    def on_monthly_report(self):
        report = get_monthly_report()
        text = format_report_text(report)

        dlg = QDialog(self)
        dlg.setWindowTitle("📊 月度经营报告")
        dlg.setMinimumSize(550, 500)
        layout = QVBoxLayout(dlg)

        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(text)
        text_edit.setObjectName("reportText")
        layout.addWidget(text_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dlg.accept)
        layout.addWidget(buttons)
        dlg.exec()

    # -------------------------------------------------------
    # Skill 6: 远程诊断助手
    # -------------------------------------------------------
    def on_remote_diagnose(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("🔧 远程诊断助手")
        dlg.setMinimumSize(650, 550)
        layout = QVBoxLayout(dlg)

        # 搜索栏
        search_row = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("输入症状关键词（如：蓝屏、开不了机、WiFi断连、风扇噪音大）")
        search_btn = QPushButton("搜索")
        search_row.addWidget(search_edit)
        search_row.addWidget(search_btn)
        layout.addLayout(search_row)

        # 快捷按钮
        quick_row = QHBoxLayout()
        for key_info in get_all_diagnose_keys():
            btn = QPushButton(key_info["title"])
            btn.setObjectName("diagnoseOptionBtn")
            btn.clicked.connect(lambda checked, k=key_info["key"]: self._show_diagnose_steps(dlg, k))
            quick_row.addWidget(btn)
        quick_row.addStretch()
        layout.addLayout(quick_row)

        # 结果区域
        result_text = QTextEdit()
        result_text.setReadOnly(True)
        result_text.setPlaceholderText("选择一个故障类型开始排查...")
        result_text.setObjectName("reportText")
        layout.addWidget(result_text)

        def do_search():
            kw = search_edit.text().strip()
            if not kw:
                return
            results = search_diagnose(kw)
            if results:
                lines = [f"找到 {len(results)} 个匹配项：\n"]
                for r in results:
                    lines.append(f"• {r['title']}（{r['match_type']}）")
                lines.append("\n点击上方快捷按钮开始排查")
                result_text.setPlainText("\n".join(lines))
            else:
                result_text.setPlainText(f"未找到与「{kw}」相关的诊断流程。\n\n当前支持的故障类型：\n" +
                                         "\n".join(f"• {k['title']}" for k in get_all_diagnose_keys()))

        search_btn.clicked.connect(do_search)
        search_edit.returnPressed.connect(do_search)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)
        dlg.exec()

    def _show_diagnose_steps(self, parent_dlg, key):
        tree = get_diagnose_tree(key)
        if not tree:
            return

        dlg = QDialog(parent_dlg)
        dlg.setWindowTitle(f"🔧 {tree['title']}")
        dlg.setMinimumSize(600, 450)
        layout = QVBoxLayout(dlg)

        title = QLabel(tree["title"])
        title.setObjectName("sectionTitleBlue")
        layout.addWidget(title)

        result_text = QTextEdit()
        result_text.setReadOnly(True)
        result_text.setObjectName("reportText")
        layout.addWidget(result_text)

        selected_path = []

        def show_step(step_idx):
            if step_idx >= len(tree["steps"]):
                return
            step = tree["steps"][step_idx]
            result_text.append(f"\n❓ {step['q']}\n")

            # 清除旧按钮
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                if item.widget() and isinstance(item.widget(), QPushButton) and item.widget().text() not in ("关闭",):
                    item.widget().deleteLater()

            for opt_name, opt_data in step["options"].items():
                btn = QPushButton(f"→ {opt_name}")
                btn.setObjectName("diagnoseOptionBtn")
                btn.clicked.connect(lambda checked, on=opt_name, od=opt_data, si=step_idx: handle_option(on, od, si))
                layout.insertWidget(layout.count() - 1, btn)

        def handle_option(opt_name, opt_data, step_idx):
            selected_path.append(opt_name)
            result_text.append(f"  ✅ {opt_name}")

            if "action" in opt_data:
                result_text.append(f"\n📋 处理方案:\n{opt_data['action']}\n")

                # 生成报告按钮
                report_btn = QPushButton("📄 生成诊断报告")
                report_btn.setObjectName("successBtn")
                report_btn.clicked.connect(lambda: self._save_diagnose_report(key, selected_path))
                layout.insertWidget(layout.count() - 1, report_btn)

            elif "next" in opt_data:
                show_step(opt_data["next"])

        show_step(0)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dlg.accept)
        layout.addWidget(close_btn)
        dlg.exec()

    def _save_diagnose_report(self, key, selected_path):
        report = generate_diagnose_report(key, selected_path)
        # 复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(report)
        QMessageBox.information(self, "诊断报告", f"报告已复制到剪贴板！\n\n{report}")

    # -------------------------------------------------------
    # Skill: 价格异动哨兵
    # -------------------------------------------------------
    def on_price_diff(self):
        """价格异动哨兵 - 在导入Word时自动对比"""
        snapshots = get_all_snapshots()
        if not snapshots:
            QMessageBox.information(self, "价格异动", "暂无历史价格快照。\n导入 Word 价格表时会自动保存快照。")
            return

        latest = get_latest_snapshot()
        if not latest:
            QMessageBox.information(self, "价格异动", "无法获取最新快照")
            return

        # 显示快照列表
        items_text = "\n".join([f"  {s['import_date']} | {s['item_count']} 条机型" for s in snapshots[:10]])
        QMessageBox.information(self, "价格快照历史", f"已有 {len(snapshots)} 个快照：\n\n{items_text}\n\n下次导入 Word 价格表时将自动对比异动。")

    # -------------------------------------------------------
    # Skill: 报价决策助手
    # -------------------------------------------------------
    def on_quote_assist(self):
        """报价决策助手"""
        dlg = QDialog(self)
        dlg.setWindowTitle("报价决策助手")
        dlg.setMinimumWidth(450)
        layout = QFormLayout(dlg)

        series_edit = QLineEdit()
        series_edit.setPlaceholderText("输入系列名称（如 Y7000P）")
        cpu_edit = QLineEdit()
        cpu_edit.setPlaceholderText("CPU（选填）")
        price_spin = QSpinBox()
        price_spin.setRange(0, 999999)
        price_spin.setPrefix("¥ ")
        customer_edit = QLineEdit()
        customer_edit.setPlaceholderText("客户名称（选填）")

        layout.addRow("系列:", series_edit)
        layout.addRow("CPU:", cpu_edit)
        layout.addRow("进货价:", price_spin)
        layout.addRow("客户:", customer_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addRow(buttons)

        if dlg.exec():
            series = series_edit.text().strip()
            if not series:
                QMessageBox.warning(self, "提示", "请输入系列名称")
                return
            suggestion = suggest_price(
                series=series,
                cpu=cpu_edit.text().strip(),
                purchase_price=price_spin.value(),
                customer_name=customer_edit.text().strip(),
            )

            conf_text = {"high": "高", "medium": "中", "low": "低"}[suggestion["confidence"]]
            text = (
                f"建议报价范围: ¥{suggestion['suggested_min']:,.0f} ~ ¥{suggestion['suggested_max']:,.0f}\n"
                f"建议中间价: ¥{suggestion['suggested_mid']:,.0f}\n"
                f"利润率: {suggestion['margin_at_mid']:.1f}%\n"
                f"置信度: {conf_text}\n\n"
                f"依据: {suggestion['basis']}"
            )
            if suggestion.get("history"):
                h = suggestion["history"]
                text += f"\n\n历史报价: {h['total_quotes']} 条\n区间: ¥{h['min_price']:,.0f} ~ ¥{h['max_price']:,.0f}\n均价: ¥{h['avg_price']:,.0f}"

            QMessageBox.information(self, "报价建议", text)

    # -------------------------------------------------------
    # Skill: 出库一条龙
    # -------------------------------------------------------
    def on_shipment_flow(self):
        """出库一条龙 - SN批量校验"""
        dlg = QDialog(self)
        dlg.setWindowTitle("出库一条龙 - SN批量校验")
        dlg.setMinimumWidth(500)
        dlg.setMinimumHeight(400)
        layout = QVBoxLayout(dlg)

        layout.addWidget(QLabel("批量输入SN（支持条码枪连续扫入，换行/逗号分隔）:"))
        sn_input = QTextEdit()
        sn_input.setPlaceholderText("扫码或粘贴SN，多个SN用换行或逗号分隔...")
        layout.addWidget(sn_input)

        qty_row = QHBoxLayout()
        qty_spin = QSpinBox()
        qty_spin.setRange(0, 9999)
        qty_spin.setValue(0)
        qty_row.addWidget(QLabel("期望数量:"))
        qty_row.addWidget(qty_spin)
        qty_row.addStretch()
        layout.addLayout(qty_row)

        result_area = QTextEdit()
        result_area.setReadOnly(True)
        layout.addWidget(result_area)

        btn_row = QHBoxLayout()
        check_btn = QPushButton("校验SN")
        def on_check():
            raw = sn_input.toPlainText()
            sn_list = parse_sn_input(raw)
            if not sn_list:
                result_area.setPlainText("未输入SN")
                return
            expected = qty_spin.value() if qty_spin.value() > 0 else None
            result = validate_sn_list(sn_list, expected)
            text = result["message"] + "\n\n"
            if result["valid"]:
                text += f"有效SN:\n" + "\n".join([f"  {sn}" for sn in result["valid"]]) + "\n"
            if result["invalid"]:
                text += f"无效SN:\n" + "\n".join([f"  {sn} - {msg}" for sn, msg in result["invalid"]]) + "\n"
            result_area.setPlainText(text)

        check_btn.clicked.connect(on_check)

        receipt_btn = QPushButton("生成出库单")
        def on_receipt():
            raw = sn_input.toPlainText()
            sn_list = parse_sn_input(raw)
            valid = [sn for sn in sn_list if validate_sn(sn)[0]]
            if not valid:
                QMessageBox.warning(self, "提示", "没有有效的SN")
                return
            receipt = generate_shipment_receipt(
                {"series": "手动出库", "customer_name": "________", "quote_price": 0, "quote_quantity": len(valid)},
                valid
            )
            clipboard = QApplication.clipboard()
            clipboard.setText(receipt)
            QMessageBox.information(self, "出库单", "出库确认单已复制到剪贴板！")

        receipt_btn.clicked.connect(on_receipt)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(check_btn)
        btn_row.addWidget(receipt_btn)
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        dlg.exec()

    # -------------------------------------------------------
    # 价格异动报告（Skill 2）
    # -------------------------------------------------------
    def _show_diff_report(self, diff):
        """弹出价格异动报告对话框"""
        dlg = QDialog(self)
        dlg.setWindowTitle("📊 价格异动报告")
        dlg.setMinimumSize(700, 500)
        layout = QVBoxLayout(dlg)

        # 标题
        title = QLabel(f"📊 价格异动报告（{diff['old_date']} → {diff['new_date']}）")
        title.setObjectName("sectionTitleBlue")
        layout.addWidget(title)

        # 摘要
        summary = QLabel(diff["summary"])
        summary.setObjectName("summaryLabel")
        layout.addWidget(summary)

        # Tab 切换
        tabs = QTabWidget()

        # === 新增 ===
        if diff["added"]:
            add_tab = QWidget()
            add_layout = QVBoxLayout(add_tab)
            add_table = QTableWidget()
            add_table.setAlternatingRowColors(True)
            add_table.setColumnCount(6)
            add_table.setHorizontalHeaderLabels(["系列", "CPU", "内存", "硬盘", "显卡", "备注"])
            add_table.setRowCount(len(diff["added"]))
            for i, item in enumerate(diff["added"]):
                add_table.setItem(i, 0, QTableWidgetItem(item.get("series", "")))
                add_table.setItem(i, 1, QTableWidgetItem(item.get("cpu", "")))
                add_table.setItem(i, 2, QTableWidgetItem(item.get("ram", "")))
                add_table.setItem(i, 3, QTableWidgetItem(item.get("storage", "")))
                add_table.setItem(i, 4, QTableWidgetItem(item.get("gpu", "")))
                add_table.setItem(i, 5, QTableWidgetItem(item.get("note", "")))
            add_table.horizontalHeader().setStretchLastSection(True)
            add_table.resizeColumnsToContents()
            add_layout.addWidget(add_table)
            tabs.addTab(add_tab, f"🔵 新增 ({len(diff['added'])})")

        # === 下架 ===
        if diff["removed"]:
            rm_tab = QWidget()
            rm_layout = QVBoxLayout(rm_tab)
            rm_label = QLabel("⚠️ 以下机型在新价格表中已下架，请检查是否有库存需要尽快出货：")
            rm_label.setObjectName("dangerSummaryLabel")
            rm_layout.addWidget(rm_label)
            rm_table = QTableWidget()
            rm_table.setAlternatingRowColors(True)
            rm_table.setColumnCount(6)
            rm_table.setHorizontalHeaderLabels(["系列", "CPU", "内存", "硬盘", "显卡", "备注"])
            rm_table.setRowCount(len(diff["removed"]))
            for i, item in enumerate(diff["removed"]):
                rm_table.setItem(i, 0, QTableWidgetItem(item.get("series", "")))
                rm_table.setItem(i, 1, QTableWidgetItem(item.get("cpu", "")))
                rm_table.setItem(i, 2, QTableWidgetItem(item.get("ram", "")))
                rm_table.setItem(i, 3, QTableWidgetItem(item.get("storage", "")))
                rm_table.setItem(i, 4, QTableWidgetItem(item.get("gpu", "")))
                rm_table.setItem(i, 5, QTableWidgetItem(item.get("note", "")))
            rm_table.horizontalHeader().setStretchLastSection(True)
            rm_table.resizeColumnsToContents()
            rm_layout.addWidget(rm_table)
            tabs.addTab(rm_tab, f"⚠️ 下架 ({len(diff['removed'])})")

        layout.addWidget(tabs)

        # 底部按钮
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dlg.accept)
        layout.addWidget(buttons)

        dlg.exec()

    # -------------------------------------------------------
    # 群发图片
    # -------------------------------------------------------
    def on_broadcast(self):
        products = get_all_products()
        if not products:
            QMessageBox.warning(self, "提示", "还没有机型数据，请先导入 Word 价格表")
            return

        # 弹出选择对话框
        dlg = QDialog(self)
        dlg.setWindowTitle("选择要群发的机型")
        dlg.setMinimumSize(600, 500)
        layout = QVBoxLayout(dlg)

        label = QLabel("勾选要生成的机型（默认全选）：")
        layout.addWidget(label)

        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(["选中", "系列", "CPU", "内存", "硬盘", "显卡", "备注"])
        table.setRowCount(len(products))
        checkboxes = []
        for i, p in enumerate(products):
            cb = QCheckBox()
            cb.setChecked(True)
            checkboxes.append(cb)
            w = QWidget()
            w_layout = QHBoxLayout(w)
            w_layout.addWidget(cb)
            w_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            w_layout.setContentsMargins(0, 0, 0, 0)
            table.setCellWidget(i, 0, w)
            table.setItem(i, 1, QTableWidgetItem(p.get("series", "")))
            table.setItem(i, 2, QTableWidgetItem(p.get("cpu", "")))
            table.setItem(i, 3, QTableWidgetItem(p.get("ram", "")))
            table.setItem(i, 4, QTableWidgetItem(p.get("storage", "")))
            table.setItem(i, 5, QTableWidgetItem(p.get("gpu", "")))
            table.setItem(i, 6, QTableWidgetItem(p.get("note", "")))

        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table)

        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("全选")
        unselect_all_btn = QPushButton("取消全选")
        select_all_btn.clicked.connect(lambda: [cb.setChecked(True) for cb in checkboxes])
        unselect_all_btn.clicked.connect(lambda: [cb.setChecked(False) for cb in checkboxes])
        btn_layout.addWidget(select_all_btn)
        btn_layout.addWidget(unselect_all_btn)
        btn_layout.addStretch()

        gen_btn = QPushButton("生成群发图片")
        gen_btn.setObjectName("successBtn")
        gen_btn.clicked.connect(dlg.accept)
        btn_layout.addWidget(gen_btn)
        layout.addLayout(btn_layout)

        if dlg.exec():
            selected = []
            for i, cb in enumerate(checkboxes):
                if cb.isChecked() and i < len(products):
                    selected.append(products[i])
            if not selected:
                QMessageBox.warning(self, "提示", "请至少选择一个机型")
                return
            paths = generate_quote_image(selected)
            msg = f"已生成 {len(paths)} 张图片:\n" + "\n".join(paths)
            QMessageBox.information(self, "生成完成", msg)

    # -------------------------------------------------------
    # 导出 Excel
    # -------------------------------------------------------
    def on_export_records_excel(self):
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        quotes = export_quotes(date_from, date_to)
        if not quotes:
            QMessageBox.warning(self, "提示", "当前筛选条件下没有报价记录")
            return
        output = export_quotes_to_excel(quotes)
        QMessageBox.information(self, "导出成功", f"报价记录已导出:\n{output}")

    def on_export_excel(self):
        """从菜单/工具栏导出全部记录"""
        self.tabs.setCurrentIndex(2)  # 切到报价记录 tab
        QMessageBox.information(self, "提示", "请切换到「报价记录」标签页，筛选后点击「导出为 Excel」")

    def on_export_json(self):
        """导出全量数据为 JSON 格式"""
        try:
            from src.utils.json_export import export_all_to_json
            from src.models import database as db
            
            output_path = export_all_to_json(db)
            QMessageBox.information(
                self, "导出成功",
                f"数据已成功导出为 JSON 格式！\n\n文件位置: {output_path}\n\n可用于数据备份或迁移到其他电脑。"
            )
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出时发生错误:\n{str(e)}")

    def on_import_json(self):
        """从 JSON 文件导入数据"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择 JSON 备份文件", "",
            "JSON 文件 (*.json);;所有文件 (*.*)"
        )
        
        if not file_path:
            return
        
        from src.utils.json_export import validate_json_file, import_from_json
        from src.models import database as db
        
        valid, message, stats = validate_json_file(file_path)
        if not valid:
            QMessageBox.warning(self, "文件无效", message)
            return
        
        reply = QMessageBox.question(
            self, "确认导入",
            f"即将导入以下数据:\n\n"
            f"• 机型: {stats.get('products', 0)} 条\n"
            f"• 批次: {stats.get('batches', 0)} 条\n"
            f"• 客户: {stats.get('customers', 0)} 条\n"
            f"• 报价: {stats.get('quotes', 0)} 条\n\n"
            f"是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            success, message, import_stats = import_from_json(file_path, db)
            
            if success:
                self.refresh_product_list()
                self.refresh_customer_list()
                self.refresh_records()
                
                QMessageBox.information(
                    self, "导入成功",
                    f"数据导入完成！\n\n"
                    f"• 机型: {import_stats.get('products', 0)} 条\n"
                    f"• 批次: {import_stats.get('batches', 0)} 条\n"
                    f"• 客户: {import_stats.get('customers', 0)} 条\n"
                    f"• 报价: {import_stats.get('quotes', 0)} 条\n\n"
                    f"注意: 如果导入的数据与现有数据重复，可能会产生重复记录。"
                )
            else:
                QMessageBox.critical(self, "导入失败", message)

    def on_show_logs(self):
        """显示操作日志"""
        dlg = OperationLogDialog(self)
        dlg.exec()


# ============================================================
# 程序入口
# ============================================================
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(APP_STYLE)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
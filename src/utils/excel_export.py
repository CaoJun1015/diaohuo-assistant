"""
Excel 导出模块：将报价记录导出为 .xlsx 格式
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_quotes_to_excel(quotes, output_path=None):
    """
    将报价记录导出为 Excel 文件。
    
    quotes: list of dict, 包含字段:
        quote_date, customer_name, series, cpu, ram, storage, gpu,
        purchase_price, quote_price, remark, paid
    
    output_path: 输出路径，默认桌面
    """
    if not output_path:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(desktop, f"报价记录_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "报价记录"

    # 表头
    headers = ["日期", "客户", "机型系列", "CPU", "内存", "硬盘", "显卡",
               "上游", "购入价", "数量", "对外报价", "毛利", "状态", "已收款", "SN", "备注", "是否打款"]
    
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据
    data_font = Font(size=10)
    data_alignment = Alignment(vertical="center")
    profit_font_good = Font(size=10, color="008000", bold=True)
    profit_font_bad = Font(size=10, color="D32F2F", bold=True)
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    total_purchase = 0
    total_quote = 0

    for row_idx, q in enumerate(quotes, 2):
        purchase_price = q.get("purchase_price", 0) or 0
        quote_price = q.get("quote_price", 0) or 0
        quantity = q.get("quote_quantity", 1) or 1
        profit = (quote_price - purchase_price) * quantity

        total_purchase += purchase_price * quantity
        total_quote += quote_price * quantity

        row_data = [
            q.get("quote_date", ""),
            q.get("customer_name", ""),
            q.get("series", ""),
            q.get("cpu", ""),
            q.get("ram", ""),
            q.get("storage", ""),
            q.get("gpu", ""),
            q.get("supplier_name", "") or "",
            purchase_price,
            quantity,
            quote_price,
            profit,
            q.get("status", "待确认"),
            q.get("received_amount", 0) or 0,
            q.get("sn_list", "") or "",
            q.get("remark", ""),
            q.get("paid", "否"),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        # 毛利列特殊着色
        profit_cell = ws.cell(row=row_idx, column=12)
        if profit > 0:
            profit_cell.font = profit_font_good
        elif profit < 0:
            profit_cell.font = profit_font_bad

        # 状态列着色
        status_cell = ws.cell(row=row_idx, column=13)
        status_colors = {
            "待确认": "9E9E9E", "已报价": "1976D2", "已出库": "F57C00",
            "已收款": "388E3C", "已取消": "BDBDBD",
        }
        status_val = q.get("status", "待确认")
        if status_val in status_colors:
            status_cell.fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
            status_cell.font = Font(size=10, color="FFFFFF", bold=True)

    # 汇总行
    summary_row = len(quotes) + 2
    ws.cell(row=summary_row, column=8, value="合计").font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=9, value=round(total_purchase, 2)).font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=11, value=round(total_quote, 2)).font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=12, value=round(total_quote - total_purchase, 2)).font = Font(bold=True, size=10, color="008000")

    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col_idx).border = thin_border

    # 列宽
    col_widths = [12, 12, 16, 14, 8, 8, 10, 10, 10, 8, 10, 10, 10, 10, 15, 20, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path